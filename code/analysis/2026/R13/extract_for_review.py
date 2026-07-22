import json
from pathlib import Path
from openpyxl import load_workbook

script_dir = Path(__file__).resolve().parent
repo_root = script_dir.parents[3]
analysis_dir = repo_root / "data/processed/analysis/2026/R13"
base_path = analysis_dir / "BookR13.xlsx"
class_path = analysis_dir / "clasificacion_R13_2026.xlsx"
out_dir = script_dir

code_questions = ["q2", "q4", "q5", "q7"]

base_wb = load_workbook(base_path, data_only=True)
ws_base = base_wb["R13"]
headers_base = [ws_base.cell(1, c).value for c in range(1, ws_base.max_column + 1)]
columns_base = {h: i + 1 for i, h in enumerate(headers_base)}

definitions = {q: [] for q in code_questions}
for r in range(2, ws_base.max_row + 1):
    q = str(ws_base.cell(r, columns_base["pregunta"]).value)
    if q in definitions:
        definitions[q].append({
            "category": str(ws_base.cell(r, columns_base["etiqueta"]).value or ""),
            "definition": str(ws_base.cell(r, columns_base["descripcion"]).value or ""),
            "examples": "",
        })

class_wb = load_workbook(class_path, data_only=True)
ws = class_wb["resultados"]
headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
columns = {h: i + 1 for i, h in enumerate(headers)}

records = []
for r in range(2, ws.max_row + 1):
    rec = {
        "row": r,
        "id": ws.cell(r, columns["ID"]).value,
        "nombre": ws.cell(r, columns["nombre"]).value,
        "segmento": ws.cell(r, columns["segmento"]).value,
        "voto2": ws.cell(r, columns["voto2"]).value,
    }
    for q in code_questions:
        rec[q] = ws.cell(r, columns[q]).value
        rec[f"codigo_{q}"] = ws.cell(r, columns[f"codigo_{q}"]).value
    records.append(rec)

out_dir.mkdir(parents=True, exist_ok=True)
(out_dir / "definitions.json").write_text(json.dumps(definitions, ensure_ascii=False, indent=2), encoding="utf-8")
(out_dir / "records.json").write_text(json.dumps(records, ensure_ascii=False, indent=2), encoding="utf-8")

print(json.dumps({
    "definition_counts": {k: len(v) for k, v in definitions.items()},
    "records": len(records),
    "headers": headers,
}, ensure_ascii=False, indent=2))
