from pathlib import Path
from openpyxl import load_workbook

repo_root = Path(__file__).resolve().parents[4]
analysis_dir = repo_root / "data/processed/analysis/2026/R13"
src = analysis_dir / "clasificacion_R13_2026.xlsx"
out = analysis_dir / "R13_codificada_corregida.xlsx"

corrections = {
    "codigo_q2": {
        2: "Lentitud e Improvisación",
        13: None,
        26: "Crítica de Identidad Política",
        37: "Seguridad Pública; Empleo y Mercado Laboral",
        39: "Lentitud e Improvisación",
        40: None,
        52: "Lentitud e Improvisación",
        55: None,
        56: None,
        68: None,
        69: None,
        77: None,
        82: None,
        83: None,
        92: None,
        96: None,
        99: None,
        106: "Economía y Costo de Vida; Reforma Previsional / AFAPs; Políticas Sociales e Inclusión",
        113: "Anticorrupción y Justicia; Políticas Sociales e Inclusión",
        116: "Políticas Sociales e Inclusión; Economía y Costo de Vida; Seguridad Pública",
    },
    "codigo_q4": {
        4: "Desencanto Político Generalizado",
        5: None,
        10: "Lentitud, Diálogo e Improvisación",
        12: "Lentitud, Diálogo e Improvisación",
        13: None,
        17: "Conflicto Previsional y Jubilatorio; Perfil Ideológico y Continuidad",
        18: None,
        19: "Herencia Adversa / Restricción Fiscal",
        32: None,
        38: None,
        39: None,
        48: None,
        49: None,
        54: None,
        55: None,
        56: "Políticas Sociales y Vulnerabilidad; Desencanto Político Generalizado",
        66: None,
        73: "Herencia Adversa / Restricción Fiscal; Lentitud, Diálogo e Improvisación",
        74: "Herencia Adversa / Restricción Fiscal; Lentitud, Diálogo e Improvisación",
        75: "Lentitud, Diálogo e Improvisación",
        77: None,
        84: "Políticas Sociales y Vulnerabilidad; Conflicto Previsional y Jubilatorio; Seguridad Pública y Violencia",
        88: "Herencia Adversa / Restricción Fiscal",
        93: None,
        105: "Atributos de Liderazgo y Tutoría; Seguridad Pública y Violencia",
        106: None,
        108: "Atributos de Liderazgo y Tutoría; Desencanto Político Generalizado",
        113: "Herencia Adversa / Restricción Fiscal",
        116: "Herencia Adversa / Restricción Fiscal",
        118: "Lentitud, Diálogo e Improvisación",
    },
    "codigo_q5": {
        3: "Shock de Seguridad y Presencia; Protección Industrial y Empleo Local; Alivio Fiscal al Trabajador y Pymes",
        5: "Políticas de Contención e Integración",
        9: "Reestructura de Gabinete y Mandato",
        12: "Redistribución y Presión al Gran Capital; Inversión Educativa y Gestión Institucional; Revisión Previsional y Edad de Retiro",
        14: "Shock de Seguridad y Presencia; Protección Industrial y Empleo Local; Inversión Educativa y Gestión Institucional",
        15: "Aceleración y Unidad del Gasto",
        16: "Políticas de Contención e Integración",
        18: None,
        20: "Aceleración y Unidad del Gasto",
        24: "Protección Industrial y Empleo Local",
        30: "Reestructura de Gabinete y Mandato; Shock de Seguridad y Presencia; Protección Industrial y Empleo Local",
        34: "Shock de Seguridad y Presencia; Políticas de Contención e Integración; Protección Industrial y Empleo Local",
        38: None,
        39: None,
        40: None,
        45: "Shock de Seguridad y Presencia",
        49: "Alivio Fiscal al Trabajador y Pymes; Protección Industrial y Empleo Local; Aceleración y Unidad del Gasto",
        53: "Políticas de Contención e Integración",
        56: None,
        57: "Reestructura de Gabinete y Mandato; Aceleración y Unidad del Gasto",
        58: "Aceleración y Unidad del Gasto",
        65: None,
        68: "Políticas de Contención e Integración; Aceleración y Unidad del Gasto",
        72: "Aceleración y Unidad del Gasto",
        75: None,
        78: "Aceleración y Unidad del Gasto; Shock de Seguridad y Presencia; Revisión Previsional y Edad de Retiro",
        83: "Aceleración y Unidad del Gasto",
        90: "Shock de Seguridad y Presencia; Reestructura de Gabinete y Mandato; Aceleración y Unidad del Gasto",
        94: "Aceleración y Unidad del Gasto; Reestructura de Gabinete y Mandato",
        96: "Aceleración y Unidad del Gasto",
        100: "Aceleración y Unidad del Gasto",
        101: "Inversión Educativa y Gestión Institucional; Protección Industrial y Empleo Local",
        106: "Aceleración y Unidad del Gasto",
        107: "Shock de Seguridad y Presencia; Políticas de Contención e Integración; Revisión Previsional y Edad de Retiro",
        110: "Alivio Fiscal al Trabajador y Pymes; Redistribución y Presión al Gran Capital",
        114: "Escucha Social Activa y Diálogo Real; Políticas de Contención e Integración",
        115: "Redistribución y Presión al Gran Capital; Políticas de Contención e Integración",
    },
    "codigo_q7": {
        3: "Continuidad Macro y Regla Fiscal",
        4: "Invisibilidad del Cambio / Estancamiento",
        8: "Invisibilidad del Cambio / Estancamiento",
        10: None,
        11: "Diferencia de Intención, Igualdad de Resultado; Continuidad Macro y Regla Fiscal",
        18: "Invisibilidad del Cambio / Estancamiento",
        20: None,
        21: "Improvisación / Barco a la Deriva; Continuidad en Presión Fiscal y Tarifas",
        27: None,
        30: "Giro Populista / Compra de Votos; Improvisación / Barco a la Deriva; Continuidad en Presión Fiscal y Tarifas",
        32: "Continuidad Macro y Regla Fiscal; Frustración de Izquierda / Desilusión",
        37: "Continuidad en Presión Fiscal y Tarifas; Invisibilidad del Cambio / Estancamiento",
        41: "Invisibilidad del Cambio / Estancamiento",
        48: None,
        49: None,
        54: None,
        55: None,
        56: None,
        58: "Invisibilidad del Cambio / Estancamiento",
        60: "Frustración de Izquierda / Desilusión",
        61: "Continuidad en Presión Fiscal y Tarifas",
        62: "Giro Asistencial y Foco en Vulnerabilidad",
        69: "Invisibilidad del Cambio / Estancamiento",
        71: None,
        77: None,
        78: "Continuidad en Presión Fiscal y Tarifas; Giro Populista / Compra de Votos; Improvisación / Barco a la Deriva",
        82: "Frustración de Izquierda / Desilusión",
        84: None,
        85: "Diferencia de Intención, Igualdad de Resultado",
        90: "Giro Populista / Compra de Votos; Improvisación / Barco a la Deriva; Continuidad en Presión Fiscal y Tarifas",
        91: "Diferencia de Intención, Igualdad de Resultado",
        96: None,
        98: "Retracción de Inversión y Empleo; Continuidad en Presión Fiscal y Tarifas",
        101: "Diferencia de Intención, Igualdad de Resultado",
        102: "Invisibilidad del Cambio / Estancamiento",
        104: "Continuidad Macro y Regla Fiscal",
        105: "Diferencia de Intención, Igualdad de Resultado",
        108: "Continuidad en Presión Fiscal y Tarifas; Improvisación / Barco a la Deriva; Invisibilidad del Cambio / Estancamiento",
        109: None,
        115: "Giro Asistencial y Foco en Vulnerabilidad; Diferencia de Intención, Igualdad de Resultado",
        118: None,
        119: "Continuidad en Presión Fiscal y Tarifas; Diferencia de Intención, Igualdad de Resultado",
    },
}

wb = load_workbook(src)
ws = wb["resultados"]
headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
columns = {h: i + 1 for i, h in enumerate(headers)}

changed = []
for header, rows in corrections.items():
    col = columns[header]
    for row, new_value in rows.items():
        cell = ws.cell(row, col)
        old_value = cell.value
        if old_value != new_value:
            cell.value = new_value
            changed.append((row, header, old_value, new_value))

out.parent.mkdir(parents=True, exist_ok=True)
wb.save(out)

print(f"saved={out}")
print(f"changes={len(changed)}")
for row, header, old_value, new_value in changed:
    print(f"{row}\t{header}\t{old_value!r}\t=>\t{new_value!r}")
