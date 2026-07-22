from __future__ import annotations

import csv
from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


YEAR = 2026
ROUND = "R7"
BASE_ANALYSIS_DIR = Path("data/processed/analysis") / str(YEAR) / ROUND
BASE_TRANSCRIPTIONS_DIR = Path("data/processed/transcriptions/output") / str(YEAR)

SEGMENTO_ORDER = [
    "Canelones",
    "Interior Coalición",
    "Interior Frente Amplio",
    "Montevideo",
]
VOTO2_ORDER = ["CM", "FA"]

HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAD3")
SECTION_FILL = PatternFill(fill_type="solid", fgColor="F3F3F3")
THIN_FONT = Font(name="Calibri", size=11)
BOLD_FONT = Font(name="Calibri", size=11, bold=True)
TITLE_FONT = Font(name="Calibri", size=13, bold=True)

QUESTION_SPECS = {
    "q10": {
        "base_file": BASE_TRANSCRIPTIONS_DIR / f"transcripcion_{ROUND}.csv",
        "output_file": BASE_ANALYSIS_DIR / f"R{ROUND}_base_de_datos_analisis_q10.xlsx",
        "layers": {
            "q10": {
                "type": "mapped",
                "order": [
                    "Sí, asistiendo",
                    "Sí, como docente o funcionario",
                    "Sí, ambos",
                    "No",
                ],
                "mapping": {
                    "A": "Sí, asistiendo",
                    "B": "Sí, como docente o funcionario",
                    "C": "Sí, ambos",
                    "D": "No",
                },
            }
        },
    },
    "q11": {
        "base_file": BASE_ANALYSIS_DIR / "mejoras_educacion_publica_q11.csv",
        "output_file": BASE_ANALYSIS_DIR / f"R{ROUND}_base_de_datos_analisis_q11.xlsx",
        "layers": {
            "codigo_q11": {
                "type": "codebook",
                "codebook": BASE_ANALYSIS_DIR / "codigos_q11.csv",
            }
        },
    },
}


def assert_exists(path: Path) -> None:
    if not path.exists():
        raise FileNotFoundError(f"No existe el archivo requerido: {path}")


def read_csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        return list(csv.DictReader(handle))


def read_code_order(path: Path) -> list[str]:
    return [row["codigo"] for row in read_csv_rows(path)]


def autosize_columns(ws) -> None:
    widths: dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            widths[cell.column] = max(widths.get(cell.column, 0), len(str(cell.value)))
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = min(width + 2, 42)


def style_header_row(ws, row_idx: int) -> None:
    for cell in ws[row_idx]:
        if cell.value is None:
            continue
        cell.font = BOLD_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_section_title(ws, row_idx: int) -> None:
    cell = ws.cell(row=row_idx, column=1)
    cell.font = TITLE_FONT
    cell.fill = SECTION_FILL


def build_counts(
    rows: list[dict[str, str]],
    layer: str,
    cross_var: str,
    category_order: list[str],
    cross_order: list[str],
) -> tuple[list[list[object]], dict[str, int], int]:
    valid_rows = [
        row for row in rows
        if row.get(layer) not in (None, "", "NA")
        and row.get(cross_var) in cross_order
    ]

    counts = {
        category: Counter(row[cross_var] for row in valid_rows if row[layer] == category)
        for category in category_order
    }
    column_totals = Counter(row[cross_var] for row in valid_rows)
    grand_total = sum(column_totals.values())

    table_rows: list[list[object]] = []
    for category in category_order:
        row_values = [category]
        row_total = 0
        for cross_value in cross_order:
            value = counts[category].get(cross_value, 0)
            row_values.append(value)
            row_total += value
        row_values.append(row_total)
        table_rows.append(row_values)

    total_row = ["Total"]
    for cross_value in cross_order:
        total_row.append(column_totals.get(cross_value, 0))
    total_row.append(grand_total)
    table_rows.append(total_row)

    return table_rows, dict(column_totals), grand_total


def build_percentages(
    count_rows: list[list[object]],
    column_totals: dict[str, int],
    cross_order: list[str],
) -> list[list[object]]:
    pct_rows: list[list[object]] = []
    for row in count_rows[:-1]:
        label = row[0]
        pct_row = [label]
        row_total = 0
        for idx, cross_value in enumerate(cross_order, start=1):
            count = row[idx]
            total = column_totals.get(cross_value, 0)
            pct_row.append((count / total) if total else 0)
            row_total += count
        grand_total = count_rows[-1][-1]
        pct_row.append((row_total / grand_total) if grand_total else 0)
        pct_rows.append(pct_row)

    total_row = ["Total"] + [1 if column_totals.get(cross_value, 0) else 0 for cross_value in cross_order] + [1]
    pct_rows.append(total_row)
    return pct_rows


def write_table(
    ws,
    start_row: int,
    title: str,
    headers: list[str],
    rows: list[list[object]],
    percent: bool = False,
) -> int:
    ws.cell(row=start_row, column=1, value=title)
    style_section_title(ws, start_row)

    header_row = start_row + 1
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col_idx, value=header)
    style_header_row(ws, header_row)

    current_row = header_row + 1
    for row in rows:
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            cell.font = THIN_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if percent and col_idx > 1:
                cell.number_format = "0.0%"
        if row[0] == "Total":
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=current_row, column=col_idx).font = BOLD_FONT
        current_row += 1

    return current_row + 1


def populate_layer_sheet(
    wb: Workbook,
    rows: list[dict[str, str]],
    layer: str,
    category_order: list[str],
    source_note: str,
) -> None:
    ws = wb.create_sheet(title=layer)
    ws.freeze_panes = "A4"

    ws["A1"] = f"Tablas de contingencia para {layer}"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Fuente base: {source_note}"

    next_row = 5
    for cross_var, cross_order in [("segmento", SEGMENTO_ORDER), ("voto2", VOTO2_ORDER)]:
        count_rows, column_totals, _ = build_counts(
            rows=rows,
            layer=layer,
            cross_var=cross_var,
            category_order=category_order,
            cross_order=cross_order,
        )
        pct_rows = build_percentages(count_rows, column_totals, cross_order)
        headers = [layer] + cross_order + ["Total"]

        next_row = write_table(
            ws=ws,
            start_row=next_row,
            title=f"Conteos por {cross_var}",
            headers=headers,
            rows=count_rows,
            percent=False,
        )
        next_row = write_table(
            ws=ws,
            start_row=next_row,
            title=f"Porcentajes por {cross_var}",
            headers=headers,
            rows=pct_rows,
            percent=True,
        )

    autosize_columns(ws)


def export_question(question_var: str) -> Path:
    spec = QUESTION_SPECS[question_var]
    base_file: Path = spec["base_file"]
    output_file: Path = spec["output_file"]

    assert_exists(base_file)
    rows = read_csv_rows(base_file)

    layers: dict[str, dict] = spec["layers"]
    prepared_rows = rows
    category_orders: dict[str, list[str]] = {}
    source_notes: dict[str, str] = {}

    for layer, layer_spec in layers.items():
        if layer_spec["type"] == "mapped":
            mapping = layer_spec["mapping"]
            for row in prepared_rows:
                raw_value = row.get(layer)
                row[layer] = mapping.get(raw_value, raw_value)
            category_orders[layer] = layer_spec["order"]
            source_notes[layer] = str(base_file)
        elif layer_spec["type"] == "codebook":
            codebook = layer_spec["codebook"]
            assert_exists(codebook)
            category_orders[layer] = read_code_order(codebook)
            source_notes[layer] = f"{base_file} | codebook: {codebook}"
        else:
            raise ValueError(f"Tipo de capa no soportado: {layer_spec['type']}")

    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "resumen"
    summary_ws["A1"] = f"Contingencias {question_var}"
    summary_ws["A1"].font = TITLE_FONT
    summary_ws["A3"] = "Base usada para cruces"
    summary_ws["B3"] = str(base_file)
    summary_ws["A5"] = "Filas base"
    summary_ws["B5"] = len(prepared_rows)

    current_row = 7
    for layer in layers:
        summary_ws.cell(row=current_row, column=1, value=f"Filas válidas {layer}")
        summary_ws.cell(
            row=current_row,
            column=2,
            value=sum(1 for row in prepared_rows if row.get(layer) not in (None, "", "NA")),
        )
        current_row += 1
        if layers[layer]["type"] == "codebook":
            summary_ws.cell(row=current_row, column=1, value=f"Codebook {layer}")
            summary_ws.cell(row=current_row, column=2, value=str(layers[layer]["codebook"]))
            current_row += 1

    autosize_columns(summary_ws)

    for layer, category_order in category_orders.items():
        populate_layer_sheet(
            wb=wb,
            rows=prepared_rows,
            layer=layer,
            category_order=category_order,
            source_note=source_notes[layer],
        )

    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_file)
    return output_file


def main() -> None:
    outputs = [export_question("q10"), export_question("q11")]
    for output in outputs:
        print(output)


if __name__ == "__main__":
    main()
