from pathlib import Path
from openpyxl import load_workbook

src = Path("data/processed/analysis/2026/R12/clasificacion_R12_2026.xlsx")
out = Path("data/processed/analysis/2026/R12/R12_codificada.xlsx")

corrections = {
    "codigo_q1": {
        17: "Igualdad y Universalidad; Dignidad y Vida",
        46: None,
    },
    "codigo_q6": {
        29: None,
        43: "Desconocimiento Total / Nulo",
        50: "Protección de Población Vulnerable",
        53: "Protección de Población Vulnerable",
        55: "Protección de Población Vulnerable",
        63: "Protección de Población Vulnerable",
        70: "Protección de Población Vulnerable",
        84: None,
    },
    "codigo_q7": {
        11: None,
        13: None,
        18: None,
        19: None,
        22: None,
        24: "Aceptación y Valoración",
        25: "Apoyo No Monetario",
        26: None,
        37: None,
        40: None,
        41: None,
        43: None,
        45: None,
        49: None,
        67: None,
        76: None,
        80: None,
        84: None,
        86: None,
        91: None,
        100: None,
        106: None,
    },
    "codigo_q8": {
        25: None,
        86: None,
    },
    "codigo_q10": {
        5: "Corresponsabilidad y Trabajo en Red; Estado como Vigilante y Corrector",
        22: None,
        36: None,
        37: None,
        54: None,
        63: None,
        86: None,
        98: "Crítica a la Gestión Estatal/Institucional",
    },
    "codigo_q12": {
        27: None,
        30: None,
        40: None,
        45: None,
        49: None,
        57: None,
        64: None,
        79: None,
        83: None,
    },
    "codigo_q14": {
        5: None,
        11: None,
        22: 'Turismo y "Círculo Virtuoso"; Sostenibilidad y Futuro; Pragmatismo Económico',
        25: None,
        36: None,
        54: None,
        86: None,
        106: None,
        111: 'Sostenibilidad y Futuro; Turismo y "Círculo Virtuoso"; Pragmatismo Económico',
    },
}

wb = load_workbook(src)
ws = wb["resultados"]
headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
columns = {h: i + 1 for i, h in enumerate(headers)}

changed = []

# Ajustes de alineación detectados en audios transcritos.
old = {h: ws.cell(35, columns[h]).value for h in ["q7", "codigo_q7", "q8", "codigo_q8", "q10", "codigo_q10", "q12", "codigo_q12", "q14", "codigo_q14"]}
shifted_35 = {
    "q7": old["q14"],
    "codigo_q7": "Oposición Ideológica / Punitivismo",
    "q8": old["q7"],
    "codigo_q8": "Pobreza Infantil y Vivienda; Crisis de Educación y Valores; Responsabilidad del Núcleo Familiar",
    "q10": old["q8"],
    "codigo_q10": "Primacía del Vínculo Biológico/Origen",
    "q12": old["q10"],
    "codigo_q12": "Industria de Celulosa y Extractivismo; Calidad y Preservación del Agua",
    "q14": old["q12"],
    "codigo_q14": "Pragmatismo Económico",
}
for header, value in shifted_35.items():
    cell = ws.cell(35, columns[header])
    changed.append((35, header, cell.value, value))
    cell.value = value

q6_46 = ws.cell(46, columns["q6"]).value
for header, value in {
    "q6": None,
    "codigo_q6": None,
    "q7": q6_46,
    "codigo_q7": "Limitación Económica Real",
}.items():
    cell = ws.cell(46, columns[header])
    changed.append((46, header, cell.value, value))
    cell.value = value

q7_93 = ws.cell(93, columns["q7"]).value
q8_93 = ws.cell(93, columns["q8"]).value
for header, value in {
    "q7": None,
    "codigo_q7": None,
    "q8": f"{q7_93} {q8_93}",
    "codigo_q8": "Pobreza Infantil y Vivienda; Responsabilidad y Garantía Estatal",
}.items():
    cell = ws.cell(93, columns[header])
    changed.append((93, header, cell.value, value))
    cell.value = value

for header, rows in corrections.items():
    col = columns[header]
    for row, new_value in rows.items():
        cell = ws.cell(row, col)
        old_value = cell.value
        if old_value != new_value:
            cell.value = new_value
            changed.append((row, header, old_value, new_value))

out.parent.mkdir(parents=True, exist_ok=True)
wb.save(out)

print(f"saved={out}")
print(f"changes={len(changed)}")
for row, header, old_value, new_value in changed:
    print(f"{row}\t{header}\t{old_value!r}\t=>\t{new_value!r}")
