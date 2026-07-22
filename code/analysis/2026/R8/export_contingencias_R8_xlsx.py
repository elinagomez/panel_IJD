from __future__ import annotations

import csv
from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


YEAR = 2026
ROUND = "R8"

BASE_ANALYSIS_DIR = Path("data/processed/analysis") / str(YEAR) / ROUND
BASE_FILE = BASE_ANALYSIS_DIR / "R8_codificada.csv"
OUTPUT_FILE = BASE_ANALYSIS_DIR / "R8_base_de_datos_analisis.xlsx"

SEGMENTO_ORDER = [
    "Canelones",
    "Interior Coalición",
    "Interior Frente Amplio",
    "Montevideo",
]
VOTO2_ORDER = ["CM", "FA"]

HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAD3")
SECTION_FILL = PatternFill(fill_type="solid", fgColor="F3F3F3")
THIN_FONT = Font(name="Calibri", size=11)
BOLD_FONT = Font(name="Calibri", size=11, bold=True)
TITLE_FONT = Font(name="Calibri", size=13, bold=True)

CLOSED_OPTION_MAPS = {
    "q3": [
        ("A", "Está actuando de manera correcta"),
        ("B", "No está actuando de manera correcta"),
        ("C", "No tengo una opinión formada"),
    ],
    "q5": [
        ("A", "Me resultaron creíbles"),
        ("B", "No me resultaron creíbles"),
        ("C", "No sabría decir"),
    ],
    "q7": [
        ("A", "en la construcción de nuevas cárceles"),
        ("B", "en la mejora de las cárceles ya existentes"),
        ("C", "tanto en la construcción de cárceles como en la mejora de las ya existentes"),
        ("D", "NO es prioritario que el Estado uruguayo invierta en cárceles."),
    ],
    "q9": [
        ("A", "Muy importante"),
        ("B", "Algo importante"),
        ("C", "Ni una cosa ni la otra"),
        ("D", "Poco importante"),
        ("E", "Nada importante"),
        ("F", "No tengo opinión"),
    ],
    "q11": [
        ("A", "Muy importante"),
        ("B", "Algo importante"),
        ("C", "Ni una cosa ni la otra"),
        ("D", "Poco importante"),
        ("E", "Nada importante"),
        ("F", "No tengo opinión"),
    ],
}

QUESTION_SPECS = {
    "q1": {
        "blocks": [
            {
                "field": "nivel_exposicion_q1",
                "type": "open",
                "title": "Nivel de exposición",
                "codebook": BASE_ANALYSIS_DIR / "codigos_q1.csv",
                "category": "Nivel de exposición",
            },
            {
                "field": "dimension_tematica_q1",
                "type": "open",
                "title": "Dimensiones temáticas",
                "codebook": BASE_ANALYSIS_DIR / "codigos_q1.csv",
                "category": "Dimensiones Temáticas (¿Qué recuerdan?)",
            },
            {
                "field": "tono_q1",
                "type": "open",
                "title": "Tono",
                "codebook": BASE_ANALYSIS_DIR / "codigos_q1.csv",
                "category": "Tono",
            },
        ]
    },
    "q2": {
        "blocks": [
            {
                "field": "postura_general_q2",
                "type": "open",
                "title": "1. Postura general",
                "codebook": BASE_ANALYSIS_DIR / "codigos_q2.csv",
                "category": "1. Postura General",
            },
            {
                "field": "percepcion_sistema_q2",
                "type": "open",
                "title": "2. Percepción del sistema",
                "codebook": BASE_ANALYSIS_DIR / "codigos_q2.csv",
                "category": "2. Percepción del Sistema",
            },
            {
                "field": "argumentos_temores_q2",
                "type": "open",
                "title": "3. Argumentos y temores",
                "codebook": BASE_ANALYSIS_DIR / "codigos_q2.csv",
                "category": "3. Argumentos y Temores",
            },
            {
                "field": "propuestas_solucion_q2",
                "type": "open",
                "title": "4. Propuestas de solución",
                "codebook": BASE_ANALYSIS_DIR / "codigos_q2.csv",
                "category": "4. Propuestas de Solución",
            },
        ]
    },
    "q3": {
        "blocks": [
            {
                "field": "q3",
                "type": "closed",
                "title": "q3",
                "options": CLOSED_OPTION_MAPS["q3"],
            }
        ]
    },
    "q4": {
        "blocks": [
             {
                 "field": "postura_debate_q4",
                 "type": "open",
                 "title": "1. Postura debate",
                 "codebook": BASE_ANALYSIS_DIR / "codigos_q4.csv",
                "category": "1. Postura Debate (Negro vs Bordaberry)",
             },
             {
                 "field": "evaluacion_oposicion_q4",
                 "type": "open",
                 "title": "2. Evaluación oposición",
                 "codebook": BASE_ANALYSIS_DIR / "codigos_q4.csv",
                "category": "2. Evaluación de la Oposición",
             },
             {
                 "field": "argumentos_justificacion_q4",
                 "type": "open",
                 "title": "3. Argumentos de justificación",
                 "codebook": BASE_ANALYSIS_DIR / "codigos_q4.csv",
                "category": "3. Argumentos (Justificación)",
             },
            {
                "field": "percepcion_sistema_q4",
                "type": "open",
                "title": "4. Percepción del sistema",
                "codebook": BASE_ANALYSIS_DIR / "codigos_q4.csv",
                "category": "4. Percepción del Sistema",
            },
        ]
    },
    "q5": {
        "blocks": [
            {
                "field": "q5",
                "type": "closed",
                "title": "q5",
                "options": CLOSED_OPTION_MAPS["q5"],
            }
        ]
    },
    "q6": {
        "blocks": [
            {
                "field": "evaluacion_q6",
                "type": "open",
                "title": "1. Evaluación",
                "codebook": BASE_ANALYSIS_DIR / "codigos_q6.csv",
                "category": "1. Evaluación",
            },
            {
                "field": "argumentos_desconfianza_q6",
                "type": "open",
                "title": "2. Argumentos de desconfianza",
                "codebook": BASE_ANALYSIS_DIR / "codigos_q6.csv",
                "category": "2. Argumentos de Desconfianza",
            },
            {
                "field": "argumentos_confianza_q6",
                "type": "open",
                "title": "3. Argumentos de confianza",
                "codebook": BASE_ANALYSIS_DIR / "codigos_q6.csv",
                "category": "3. Argumentos de Confianza",
            },
            {
                "field": "influencia_externa_q6",
                "type": "open",
                "title": "4. Influencia externa",
                "codebook": BASE_ANALYSIS_DIR / "codigos_q6.csv",
                "category": "4. Influencia Externa",
            },
            {
                "field": "contexto_otros_q6",
                "type": "open",
                "title": "5. Contexto / otros",
                "codebook": BASE_ANALYSIS_DIR / "codigos_q6.csv",
                "category": "5. Contexto / Otros",
            },
        ]
    },
    "q7": {
        "blocks": [
            {
                "field": "q7",
                "type": "closed",
                "title": "q7",
                "options": CLOSED_OPTION_MAPS["q7"],
            }
        ]
    },
    "q8": {
        "blocks": [
            {
                "field": "prioridad_inversion_q8",
                "type": "open",
                "title": "1. Prioridad de inversión",
                "codebook": BASE_ANALYSIS_DIR / "codigos_q8.csv",
                "category": "1. Prioridad de Inversión",
            },
            {
                "field": "justificacion_rehabilitacion_q8",
                "type": "open",
                "title": "2. Justificación: rehabilitación",
                "codebook": BASE_ANALYSIS_DIR / "codigos_q8.csv",
                "category": "2. Justificación: Rehabilitación",
            },
            {
                "field": "justificacion_punitivismo_q8",
                "type": "open",
                "title": "3. Justificación: punitivismo",
                "codebook": BASE_ANALYSIS_DIR / "codigos_q8.csv",
                "category": "3. Justificación: Punitivismo",
            },
            {
                "field": "gestion_recursos_q8",
                "type": "open",
                "title": "4. Gestión y recursos",
                "codebook": BASE_ANALYSIS_DIR / "codigos_q8.csv",
                "category": "4. Gestión y Recursos",
            },
        ]
    },
    "q9": {
        "blocks": [
            {
                "field": "q9",
                "type": "closed",
                "title": "q9",
                "options": CLOSED_OPTION_MAPS["q9"],
            }
        ]
    },
    "q10": {
        "blocks": [
            {
                "field": "codigo_q10",
                "type": "open",
                "title": "q10",
                "codebook": BASE_ANALYSIS_DIR / "codigos_q10.csv",
                "category": "Única",
            }
        ]
    },
    "q11": {
        "blocks": [
            {
                "field": "q11",
                "type": "closed",
                "title": "q11",
                "options": CLOSED_OPTION_MAPS["q11"],
            }
        ]
    },
    "q12": {
        "blocks": [
            {
                "field": "codigo_q12",
                "type": "open",
                "title": "q12",
                "codebook": BASE_ANALYSIS_DIR / "codigos_q12.csv",
                "category": "Única",
            }
        ]
    },
}


def assert_exists(path: Path) -> None:
    if not path.exists():
        raise FileNotFoundError(f"No existe el archivo requerido: {path}")


def read_csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        return list(csv.DictReader(handle))


def is_missing(value: object) -> bool:
    if value is None:
        return True
    text = str(value).strip()
    return text in {"", "NA", "N/A", "None"}


def style_header_row(ws, row_idx: int, start_col: int, end_col: int) -> None:
    for col_idx in range(start_col, end_col + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        if cell.value is None:
            continue
        cell.font = BOLD_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_section_title(ws, row_idx: int, col_idx: int = 1) -> None:
    cell = ws.cell(row=row_idx, column=col_idx)
    cell.font = TITLE_FONT
    cell.fill = SECTION_FILL
    cell.alignment = Alignment(vertical="center")


def autosize_columns(ws, max_width: int = 42) -> None:
    widths: dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            widths[cell.column] = max(widths.get(cell.column, 0), len(str(cell.value)))
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = min(width + 2, max_width)


def load_open_codebook(path: Path, category: str) -> tuple[list[str], list[dict[str, str]]]:
    assert_exists(path)
    rows = read_csv_rows(path)
    filtered = [row for row in rows if row.get("categoria") == category]
    if not filtered:
        raise ValueError(f"No se encontraron filas para la categoría '{category}' en {path}")

    order = [row["codigo"] for row in filtered]
    codebook_rows = [
        {
            "code": row["codigo"],
            "label": row["codigo"],
            "definition": row["descripcion"],
        }
        for row in filtered
    ]
    return order, codebook_rows


def load_closed_codebook(options: list[tuple[str, str]]) -> tuple[list[str], list[dict[str, str]]]:
    order = [label for _, label in options]
    codebook_rows = [{"code": code, "label": label} for code, label in options]
    return order, codebook_rows


def ensure_expected_columns(rows: list[dict[str, str]], columns: list[str]) -> None:
    available = set(rows[0].keys()) if rows else set()
    missing = [column for column in columns if column not in available]
    if missing:
        raise ValueError(f"Faltan columnas requeridas en la base: {', '.join(missing)}")


def validate_values(rows: list[dict[str, str]], field: str, expected: list[str]) -> None:
    expected_set = set(expected)
    unexpected = sorted(
        {
            str(row.get(field)).strip()
            for row in rows
            if not is_missing(row.get(field)) and str(row.get(field)).strip() not in expected_set
        }
    )
    if unexpected:
        raise ValueError(
            f"La columna {field} contiene valores fuera del orden esperado: {', '.join(unexpected)}"
        )


def build_counts(
    rows: list[dict[str, str]],
    field: str,
    category_order: list[str],
) -> tuple[list[list[object]], dict[str, int], dict[str, int], int]:
    valid_rows = [row for row in rows if not is_missing(row.get(field))]

    segmento_totals = Counter(
        row["segmento"]
        for row in valid_rows
        if row.get("segmento") in SEGMENTO_ORDER
    )
    voto2_totals = Counter(
        row["voto2"]
        for row in valid_rows
        if row.get("voto2") in VOTO2_ORDER
    )
    grand_total = len(valid_rows)

    count_rows: list[list[object]] = []
    for category in category_order:
        matching = [row for row in valid_rows if row.get(field) == category]
        row_values: list[object] = [category]
        for segmento in SEGMENTO_ORDER:
            row_values.append(sum(1 for row in matching if row.get("segmento") == segmento))
        for voto2 in VOTO2_ORDER:
            row_values.append(sum(1 for row in matching if row.get("voto2") == voto2))
        row_values.append(len(matching))
        count_rows.append(row_values)

    total_row: list[object] = ["Total"]
    total_row.extend(segmento_totals.get(segmento, 0) for segmento in SEGMENTO_ORDER)
    total_row.extend(voto2_totals.get(voto2, 0) for voto2 in VOTO2_ORDER)
    total_row.append(grand_total)
    count_rows.append(total_row)

    return count_rows, dict(segmento_totals), dict(voto2_totals), grand_total


def build_percentages(
    count_rows: list[list[object]],
    segmento_totals: dict[str, int],
    voto2_totals: dict[str, int],
    grand_total: int,
) -> list[list[object]]:
    pct_rows: list[list[object]] = []

    for row in count_rows[:-1]:
        pct_row: list[object] = [row[0]]
        for idx, segmento in enumerate(SEGMENTO_ORDER, start=1):
            total = segmento_totals.get(segmento, 0)
            pct_row.append((row[idx] / total) if total else 0)
        voto_start = 1 + len(SEGMENTO_ORDER)
        for offset, voto2 in enumerate(VOTO2_ORDER, start=voto_start):
            total = voto2_totals.get(voto2, 0)
            pct_row.append((row[offset] / total) if total else 0)
        pct_row.append((row[-1] / grand_total) if grand_total else 0)
        pct_rows.append(pct_row)

    total_row: list[object] = ["Total"]
    total_row.extend(1 if segmento_totals.get(segmento, 0) else 0 for segmento in SEGMENTO_ORDER)
    total_row.extend(1 if voto2_totals.get(voto2, 0) else 0 for voto2 in VOTO2_ORDER)
    total_row.append(1 if grand_total else 0)
    pct_rows.append(total_row)
    return pct_rows


def write_cross_table(
    ws,
    start_row: int,
    field_label: str,
    title: str,
    rows: list[list[object]],
    percent: bool,
) -> int:
    ws.cell(row=start_row, column=1, value=title)
    style_section_title(ws, start_row)

    top_header_row = start_row + 2
    ws.cell(row=top_header_row, column=1, value=field_label)
    ws.cell(row=top_header_row, column=2, value="segmento")
    ws.cell(row=top_header_row, column=6, value="voto2")
    ws.cell(row=top_header_row, column=8, value="Total")
    style_header_row(ws, top_header_row, 1, 8)

    second_header_row = top_header_row + 1
    headers = ["Etiqueta"] + SEGMENTO_ORDER + VOTO2_ORDER + ["Total"]
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=second_header_row, column=col_idx, value=header)
    style_header_row(ws, second_header_row, 1, 8)

    current_row = second_header_row + 1
    for row in rows:
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            cell.font = THIN_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if percent and col_idx > 1:
                cell.number_format = "0.0%"
        if row[0] == "Total":
            for col_idx in range(1, 9):
                ws.cell(row=current_row, column=col_idx).font = BOLD_FONT
        current_row += 1

    return current_row


def write_codebook_table(
    ws,
    start_row: int,
    start_col: int,
    block_type: str,
    block_title: str,
    rows: list[dict[str, str]],
) -> int:
    ws.cell(row=start_row, column=start_col, value=f"Códigos - {block_title}")
    style_section_title(ws, start_row, start_col)

    header_row = start_row + 1
    if block_type == "open":
        headers = ["Código", "Etiqueta", "Definición"]
    else:
        headers = ["Código", "Etiqueta"]

    for offset, header in enumerate(headers):
        ws.cell(row=header_row, column=start_col + offset, value=header)
    style_header_row(ws, header_row, start_col, start_col + len(headers) - 1)

    current_row = header_row + 1
    for row in rows:
        values = [row["code"], row["label"]]
        if block_type == "open":
            values.append(row["definition"])
        for offset, value in enumerate(values):
            cell = ws.cell(row=current_row, column=start_col + offset, value=value)
            cell.font = THIN_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        current_row += 1

    return current_row


def write_question_sheet(
    wb: Workbook,
    question: str,
    rows: list[dict[str, str]],
    summary_rows: list[dict[str, object]],
) -> None:
    spec = QUESTION_SPECS[question]
    ws = wb.create_sheet(title=question)
    ws.freeze_panes = "A5"

    ws["A1"] = f"{ROUND} {YEAR} - {question}"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Base usada para cruces: {BASE_FILE}"
    ws["A3"] = f"Casos en base: {len(rows)}"

    start_row = 5
    for block in spec["blocks"]:
        if block["type"] == "open":
            category_order, codebook_rows = load_open_codebook(block["codebook"], block["category"])
            source_note = str(block["codebook"])
        else:
            category_order, codebook_rows = load_closed_codebook(block["options"])
            source_note = "option_maps de consolidar_R8_codificada.R"

        validate_values(rows, block["field"], category_order)

        count_rows, segmento_totals, voto2_totals, grand_total = build_counts(
            rows=rows,
            field=block["field"],
            category_order=category_order,
        )
        pct_rows = build_percentages(count_rows, segmento_totals, voto2_totals, grand_total)

        ws.cell(row=start_row - 1, column=1, value=f"Fuente bloque: {source_note}")
        ws.cell(row=start_row - 1, column=1).font = THIN_FONT

        counts_end = write_cross_table(
            ws=ws,
            start_row=start_row,
            field_label=block["field"],
            title=f"{block['title']} - Conteos",
            rows=count_rows,
            percent=False,
        )
        pct_end = write_cross_table(
            ws=ws,
            start_row=counts_end + 1,
            field_label=block["field"],
            title=f"{block['title']} - Porcentajes por columna",
            rows=pct_rows,
            percent=True,
        )
        codebook_end = write_codebook_table(
            ws=ws,
            start_row=start_row,
            start_col=10,
            block_type=block["type"],
            block_title=block["title"],
            rows=codebook_rows,
        )

        summary_rows.append(
            {
                "pregunta": question,
                "variable": block["field"],
                "tipo": block["type"],
                "filas_validas": grand_total,
                "fuente_codigos": source_note,
            }
        )

        start_row = max(pct_end, codebook_end) + 3

    ws.column_dimensions["A"].width = 34
    for column in ["B", "C", "D", "E"]:
        ws.column_dimensions[column].width = 18
    for column in ["F", "G", "H"]:
        ws.column_dimensions[column].width = 14
    ws.column_dimensions["I"].width = 3
    ws.column_dimensions["J"].width = 24
    ws.column_dimensions["K"].width = 28
    ws.column_dimensions["L"].width = 90


def build_summary_sheet(
    wb: Workbook,
    rows: list[dict[str, str]],
    summary_rows: list[dict[str, object]],
) -> None:
    ws = wb.active
    ws.title = "resumen"

    ws["A1"] = f"Contingencias {ROUND} {YEAR}"
    ws["A1"].font = TITLE_FONT
    ws["A3"] = "Base usada"
    ws["B3"] = str(BASE_FILE)
    ws["A4"] = "Archivo de salida"
    ws["B4"] = str(OUTPUT_FILE)
    ws["A5"] = "Casos en base"
    ws["B5"] = len(rows)
    ws["A7"] = "Detalle de bloques"
    ws["A7"].font = BOLD_FONT

    headers = ["pregunta", "variable", "tipo", "filas_validas", "fuente_codigos"]
    header_row = 8
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col_idx, value=header)
    style_header_row(ws, header_row, 1, len(headers))

    current_row = header_row + 1
    for row in summary_rows:
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=row[header])
            cell.font = THIN_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        current_row += 1

    autosize_columns(ws, max_width=60)


def main() -> None:
    assert_exists(BASE_FILE)
    rows = read_csv_rows(BASE_FILE)

    required_columns = ["segmento", "voto2"]
    for question_spec in QUESTION_SPECS.values():
        required_columns.extend(block["field"] for block in question_spec["blocks"])
        for block in question_spec["blocks"]:
            if block["type"] == "open":
                assert_exists(block["codebook"])

    ensure_expected_columns(rows, sorted(set(required_columns)))

    summary_rows: list[dict[str, object]] = []
    wb = Workbook()

    for question in QUESTION_SPECS:
        write_question_sheet(wb, question, rows, summary_rows)

    build_summary_sheet(wb, rows, summary_rows)

    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_FILE)
    print(OUTPUT_FILE)


if __name__ == "__main__":
    main()
