from __future__ import annotations

import argparse
import csv
import json
import os
import sys
import textwrap
import time
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openai import OpenAI
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


REPO_ROOT = Path(__file__).resolve().parents[4]
ANALYSIS_DIR = REPO_ROOT / "data" / "processed" / "analysis" / "2026" / "R8"
TRANSCRIPTION_FILE = (
    REPO_ROOT
    / "data"
    / "processed"
    / "transcriptions"
    / "output"
    / "2026"
    / "transcripcion_R8.csv"
)

AUDIT_WORKBOOK = ANALYSIS_DIR / "R8_auditoria_codificacion.xlsx"
AUDIT_LONG_CSV = ANALYSIS_DIR / "R8_auditoria_long.csv"
REVIEWED_WIDE_WORKBOOK = ANALYSIS_DIR / "R8_codificada_revisada.xlsx"
REVIEW_MEMO = ANALYSIS_DIR / "R8_hallazgos_revision.md"
CACHE_FILE = ANALYSIS_DIR / "R8_revision_cache.jsonl"

CURRENT_TIMESTAMP = datetime.now(timezone.utc).astimezone().strftime("%Y-%m-%d %H:%M:%S %Z")
CURRENT_DATE = datetime.now(timezone.utc).astimezone().strftime("%Y-%m-%d")

MODEL_DEFAULT = "gpt-5.4-mini"

REVIEW_STATES = [
    "MANTENER",
    "CORREGIR",
    "MISSING_VALIDO",
    "SIN_ENCAJE_CODEBOOK",
]

PROBLEM_TYPES = [
    "NINGUNO",
    "SUBCODIFICACION",
    "SOBRECODIFICACION",
    "SESGO_DE_DEFAULT",
    "ERROR_PIPELINE",
]

OPTION_MAPS = {
    "q3": {
        "A": "Está actuando de manera correcta",
        "B": "No está actuando de manera correcta",
        "C": "No tengo una opinión formada",
    },
    "q5": {
        "A": "Me resultaron creíbles",
        "B": "No me resultaron creíbles",
        "C": "No sabría decir",
    },
    "q7": {
        "A": "en la construcción de nuevas cárceles",
        "B": "en la mejora de las cárceles ya existentes",
        "C": "tanto en la construcción de cárceles como en la mejora de las ya existentes",
        "D": "NO es prioritario que el Estado uruguayo invierta en cárceles.",
    },
    "q9": {
        "A": "Muy importante",
        "B": "Algo importante",
        "C": "Ni una cosa ni la otra",
        "D": "Poco importante",
        "E": "Nada importante",
        "F": "No tengo opinión",
    },
    "q11": {
        "A": "Muy importante",
        "B": "Algo importante",
        "C": "Ni una cosa ni la otra",
        "D": "Poco importante",
        "E": "Nada importante",
        "F": "No tengo opinión",
    },
}

SYSTEMIC_DIMENSIONS_OF_INTEREST = {
    ("q2", "percepcion_sistema_q2"),
    ("q4", "argumentos_justificacion_q4"),
    ("q6", "influencia_externa_q6"),
    ("q1", "dimension_tematica_q1"),
    ("q1", "tono_q1"),
}

BASIC_METADATA = [
    "nombre",
    "edad",
    "genero",
    "numero",
    "departamento",
    "n_educativo",
    "oficio",
    "voto",
    "segmento",
    "voto2",
    "etiqueta",
]


@dataclass(frozen=True)
class DimensionSpec:
    category: str
    field: str
    codes: list[dict[str, str]]


@dataclass(frozen=True)
class QuestionSpec:
    question: str
    tema: str
    open_col: str
    open_question_text: str
    output_file: Path
    codebook_file: Path
    category_to_field: dict[str, str]
    closed_col: str | None = None
    closed_question_text: str | None = None
    closed_value_map: dict[str, str] | None = None


def normalize_space(value: str | None) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\r", " ").replace("\n", " ").split())


def clean_code_value(value: str | None) -> str | None:
    text = normalize_space(value)
    if not text:
        return None
    if text.upper() in {"NA", "N/A", "NULL"}:
        return None
    return text


def text_is_missing(value: str | None) -> bool:
    text = normalize_space(value)
    if not text:
        return True
    lowered = text.lower()
    return lowered in {"na", "n/a", "...", ".", "..", "-", "--"}


def recode_option_letter(question: str, value: str | None) -> str | None:
    cleaned = clean_code_value(value)
    if cleaned is None:
        return None
    mapping = OPTION_MAPS.get(question)
    if mapping is None:
        return cleaned
    return mapping.get(cleaned, cleaned)


def load_csv(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        return list(csv.DictReader(handle))


def write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def load_codebook_dimensions(path: Path, category_to_field: dict[str, str]) -> list[DimensionSpec]:
    rows = load_csv(path)
    grouped: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in rows:
        grouped[row["categoria"]].append(
            {
                "codigo": normalize_space(row["codigo"]),
                "descripcion": normalize_space(row["descripcion"]),
            }
        )

    dimensions: list[DimensionSpec] = []
    for category, field in category_to_field.items():
        if category not in grouped:
            raise RuntimeError(f"No se encontró la categoría '{category}' en {path.name}.")
        codes = []
        for index, code in enumerate(grouped[category], start=1):
            codes.append(
                {
                    "code_id": f"OPT_{index:02d}",
                    "codigo": code["codigo"],
                    "descripcion": code["descripcion"],
                }
            )
        dimensions.append(DimensionSpec(category=category, field=field, codes=codes))
    return dimensions


def question_specs() -> list[QuestionSpec]:
    return [
        QuestionSpec(
            question="q1",
            tema="interpelacion_recuerdos",
            open_col="q1",
            open_question_text=(
                "En estos últimos días se produjo una interpelación del Ministro del Interior Carlos Negro. "
                "¿Viste o escuchaste algo sobre este tema? ¿Qué recuerdas?"
            ),
            output_file=ANALYSIS_DIR / "interpelacion_recuerdos_q1.csv",
            codebook_file=ANALYSIS_DIR / "codigos_q1.csv",
            category_to_field={
                "Nivel de exposición": "nivel_exposicion_q1",
                "Dimensiones Temáticas (¿Qué recuerdan?)": "dimension_tematica_q1",
                "Tono": "tono_q1",
            },
        ),
        QuestionSpec(
            question="q2",
            tema="debate_liberar_presos",
            open_col="q2",
            open_question_text=(
                "Uno de los puntos álgidos del debate fue cuando Bordaberry acusó al gobierno de pretender "
                "liberar presos con el nuevo Plan Nacional de Seguridad Pública. El Ministro Negro defendió "
                "la propuesta aclarando que lo que se busca es que NO ingresen a la cárcel personas con delitos "
                "de muy leve entidad para evitar que sean reclutados para formar parte de una banda criminal. "
                "¿Cuál es tu opinión sobre este debate?"
            ),
            output_file=ANALYSIS_DIR / "debate_liberar_presos_q2.csv",
            codebook_file=ANALYSIS_DIR / "codigos_q2.csv",
            category_to_field={
                "1. Postura General": "postura_general_q2",
                "2. Percepción del Sistema": "percepcion_sistema_q2",
                "3. Argumentos y Temores": "argumentos_temores_q2",
                "4. Propuestas de Solución": "propuestas_solucion_q2",
            },
        ),
        QuestionSpec(
            question="q4",
            tema="accionar_oposicion",
            open_col="q4",
            open_question_text="¿Por qué?",
            output_file=ANALYSIS_DIR / "accionar_oposicion_q4.csv",
            codebook_file=ANALYSIS_DIR / "codigos_q4.csv",
            category_to_field={
                "1. Postura Debate (Negro vs Bordaberry)": "postura_debate_q4",
                "2. Evaluación de la Oposición": "evaluacion_oposicion_q4",
                "3. Argumentos (Justificación)": "argumentos_justificacion_q4",
                "4. Percepción del Sistema": "percepcion_sistema_q4",
            },
            closed_col="q3",
            closed_question_text=(
                "¿Cómo evalúas el accionar de la oposición en este caso? "
                "Opciones: Está actuando de manera correcta; No está actuando de manera correcta; "
                "No tengo una opinión formada."
            ),
            closed_value_map=OPTION_MAPS["q3"],
        ),
        QuestionSpec(
            question="q6",
            tema="credibilidad_cifras",
            open_col="q6",
            open_question_text="¿Por qué?",
            output_file=ANALYSIS_DIR / "credibilidad_cifras_q6.csv",
            codebook_file=ANALYSIS_DIR / "codigos_q6.csv",
            category_to_field={
                "1. Evaluación": "evaluacion_q6",
                "2. Argumentos de Desconfianza": "argumentos_desconfianza_q6",
                "3. Argumentos de Confianza": "argumentos_confianza_q6",
                "4. Influencia Externa": "influencia_externa_q6",
                "5. Contexto / Otros": "contexto_otros_q6",
            },
            closed_col="q5",
            closed_question_text=(
                "Durante la interpelación, el Ministro del Interior comunicó cifras que muestran descensos "
                "en homicidios, hurtos, rapiñas, estafas informáticas y abigeato. "
                "¿Cuán creíble te resultan las cifras comunicadas por el Ministerio del Interior? "
                "Opciones: Me resultaron creíbles; No me resultaron creíbles; No sabría decir."
            ),
            closed_value_map=OPTION_MAPS["q5"],
        ),
        QuestionSpec(
            question="q8",
            tema="inversion_carceles",
            open_col="q8",
            open_question_text="¿Por qué?",
            output_file=ANALYSIS_DIR / "inversion_carceles_q8.csv",
            codebook_file=ANALYSIS_DIR / "codigos_q8.csv",
            category_to_field={
                "1. Prioridad de Inversión": "prioridad_inversion_q8",
                "2. Justificación: Rehabilitación": "justificacion_rehabilitacion_q8",
                "3. Justificación: Punitivismo": "justificacion_punitivismo_q8",
                "4. Gestión y Recursos": "gestion_recursos_q8",
            },
            closed_col="q7",
            closed_question_text=(
                "Actualmente en Uruguay hay más de 16.000 personas privadas de libertad en cárceles. "
                "Como informan diversos medios hay varias cárceles con sobrepoblación y otras con problemas "
                "para atender a los internos. ¿Cuál de las siguientes frases representa mejor tu opinión "
                "sobre las responsabilidades del Estado uruguayo? ¿Crees que es prioritario invertir: "
                "en la construcción de nuevas cárceles; en la mejora de las cárceles ya existentes; "
                "tanto en la construcción de cárceles como en la mejora de las ya existentes; "
                "o NO es prioritario que el Estado uruguayo invierta en cárceles?"
            ),
            closed_value_map=OPTION_MAPS["q7"],
        ),
        QuestionSpec(
            question="q10",
            tema="importancia_rehabilitacion",
            open_col="q10",
            open_question_text="¿Qué te hace pensar así?",
            output_file=ANALYSIS_DIR / "importancia_rehabilitacion_q10.csv",
            codebook_file=ANALYSIS_DIR / "codigos_q10.csv",
            category_to_field={"Única": "codigo_q10"},
            closed_col="q9",
            closed_question_text=(
                "¿Cuán importante te parece que son las políticas de rehabilitación para las personas "
                "que están en la cárcel? Opciones: Muy importante; Algo importante; Ni una cosa ni la otra; "
                "Poco importante; Nada importante; No tengo opinión."
            ),
            closed_value_map=OPTION_MAPS["q9"],
        ),
        QuestionSpec(
            question="q12",
            tema="trabajo_estudio_carcel",
            open_col="q12",
            open_question_text="¿Qué te hace pensar así?",
            output_file=ANALYSIS_DIR / "trabajo_estudio_carcel_q12.csv",
            codebook_file=ANALYSIS_DIR / "codigos_q12.csv",
            category_to_field={"Única": "codigo_q12"},
            closed_col="q11",
            closed_question_text=(
                "¿Cuán importante te parece que se ofrezca la posibilidad de trabajar y estudiar a las "
                "personas que están en la cárcel? Opciones: Muy importante; Algo importante; Ni una cosa ni la otra; "
                "Poco importante; Nada importante; No tengo opinión."
            ),
            closed_value_map=OPTION_MAPS["q11"],
        ),
    ]


def enrich_closed_answers(rows: list[dict[str, str]]) -> None:
    for row in rows:
        for question, mapping in OPTION_MAPS.items():
            if question in row:
                cleaned = clean_code_value(row.get(question))
                if cleaned is None:
                    row[f"{question}_texto"] = None
                else:
                    row[f"{question}_texto"] = mapping.get(cleaned, cleaned)


def load_current_outputs(specs: list[QuestionSpec]) -> dict[tuple[str, str], dict[str, str | None]]:
    current: dict[tuple[str, str], dict[str, str | None]] = {}
    for spec in specs:
        dimensions = load_codebook_dimensions(spec.codebook_file, spec.category_to_field)
        output_rows = load_csv(spec.output_file)
        for row in output_rows:
            key = (spec.question, row["numero"])
            current[key] = {}
            for dimension in dimensions:
                current[key][dimension.field] = clean_code_value(row.get(dimension.field))
    return current


def compute_systemic_alerts(
    rows: list[dict[str, str]],
    specs: list[QuestionSpec],
    current_outputs: dict[tuple[str, str], dict[str, str | None]],
) -> dict[tuple[str, str], dict[str, Any]]:
    alerts: dict[tuple[str, str], dict[str, Any]] = {}
    for spec in specs:
        dimensions = load_codebook_dimensions(spec.codebook_file, spec.category_to_field)
        for dimension in dimensions:
            values: list[str] = []
            usable_text_with_missing = 0
            total = 0
            for row in rows:
                total += 1
                current_code = current_outputs[(spec.question, row["numero"])].get(dimension.field)
                if current_code is None:
                    values.append("__NA__")
                    if not text_is_missing(row.get(spec.open_col)):
                        usable_text_with_missing += 1
                else:
                    values.append(current_code)
            counts = Counter(values)
            top_code, top_count = counts.most_common(1)[0]
            na_count = counts.get("__NA__", 0)
            top_share = top_count / total
            na_share = na_count / total
            alert_parts: list[str] = []
            if na_share >= 0.75:
                alert_parts.append(
                    f"Concentración en NA: {na_count}/{total}; con texto utilizable en {usable_text_with_missing}/{total}."
                )
            if top_code != "__NA__" and top_share >= 0.85:
                alert_parts.append(
                    f"Código dominante: '{top_code}' en {top_count}/{total} ({top_share:.1%})."
                )
            if (
                (spec.question, dimension.field) in SYSTEMIC_DIMENSIONS_OF_INTEREST
                and top_code != "__NA__"
                and top_share >= 0.75
                and not any(part.startswith("Código dominante:") for part in alert_parts)
            ):
                alert_parts.append(
                    f"Código dominante: '{top_code}' en {top_count}/{total} ({top_share:.1%})."
                )
            alerts[(spec.question, dimension.field)] = {
                "question": spec.question,
                "field": dimension.field,
                "dimension": dimension.category,
                "total": total,
                "na_count": na_count,
                "top_code": None if top_code == "__NA__" else top_code,
                "top_count": 0 if top_code == "__NA__" else top_count,
                "top_share": 0 if top_code == "__NA__" else top_share,
                "alert_text": " ".join(alert_parts).strip(),
            }
    return alerts


def load_cache() -> dict[str, dict[str, Any]]:
    if not CACHE_FILE.exists():
        return {}
    cached: dict[str, dict[str, Any]] = {}
    with CACHE_FILE.open(encoding="utf-8") as handle:
        for line in handle:
            line = line.strip()
            if not line:
                continue
            record = json.loads(line)
            cached[record["key"]] = record
    return cached


def append_cache(record: dict[str, Any]) -> None:
    with CACHE_FILE.open("a", encoding="utf-8") as handle:
        handle.write(json.dumps(record, ensure_ascii=False) + "\n")


def build_prompt_payload(
    row: dict[str, str],
    spec: QuestionSpec,
    dimensions: list[DimensionSpec],
    current_codes: dict[str, str | None],
    systemic_alerts: dict[tuple[str, str], dict[str, Any]],
) -> str:
    dimension_blocks = []
    for dimension in dimensions:
        allowed_codes = "\n".join(
            f"- {code['code_id']} => {code['codigo']}: {code['descripcion']}" for code in dimension.codes
        )
        alert = systemic_alerts[(spec.question, dimension.field)]["alert_text"] or "Sin alerta sistémica especial."
        current = current_codes.get(dimension.field) or "NA"
        dimension_blocks.append(
            textwrap.dedent(
                f"""
                DIMENSIÓN: {dimension.category}
                Campo de salida: {dimension.field}
                Código actual: {current}
                Alerta sistémica: {alert}
                Códigos permitidos:
                {allowed_codes}
                """
            ).strip()
        )

    closed_context = ""
    if spec.closed_col and spec.closed_question_text:
        closed_answer = row.get(f"{spec.closed_col}_texto") or "NA"
        closed_context = textwrap.dedent(
            f"""
            Pregunta cerrada previa:
            - Columna: {spec.closed_col}
            - Texto: {spec.closed_question_text}
            - Respuesta registrada: {closed_answer}
            """
        ).strip()

    open_text = row.get(spec.open_col) or ""
    return textwrap.dedent(
        f"""
        Revisa una respuesta abierta de encuesta ya codificada. Debes auditar cada dimensión de forma independiente.

        Reglas:
        1. Usa solo estos estados: MANTENER, CORREGIR, MISSING_VALIDO, SIN_ENCAJE_CODEBOOK.
        2. Usa solo estos tipos de problema: NINGUNO, SUBCODIFICACION, SOBRECODIFICACION, SESGO_DE_DEFAULT, ERROR_PIPELINE.
        3. Si el texto abierto está vacío, es literal "...", o no tiene contenido interpretable, usa MISSING_VALIDO y codigo_revisado = null.
        4. Si el código actual está vacío/NA pero el verbatim sí activa un código claro, corrige y usa SUBCODIFICACION o ERROR_PIPELINE según corresponda.
        5. Si el código actual parece un comodín dominante sin base textual suficiente, corrige y usa SESGO_DE_DEFAULT.
                6. Si la respuesta es sustantiva pero no encaja limpiamente en ningún código vigente, usa SIN_ENCAJE_CODEBOOK, codigo_revisado = null y requiere_cambio_codebook = true.
        7. Si mantienes el código actual, codigo_revisado debe ser el ID interno del código actual.
        8. Si corriges, codigo_revisado debe ser uno de los códigos permitidos de esa dimensión o null cuando corresponda eliminar una sobrecodificación.
        9. La justificación debe ser breve, concreta y basada en evidencia textual.
        10. La respuesta cerrada previa, si existe, sirve solo como contexto. No reemplaza al verbatim abierto.

        Metadatos del caso:
        - numero: {row['numero']}
        - nombre: {row.get('nombre') or 'NA'}
        - edad: {row.get('edad') or 'NA'}
        - genero: {row.get('genero') or 'NA'}
        - departamento: {row.get('departamento') or 'NA'}
        - pregunta abierta: {spec.question}
        - texto de la pregunta: {spec.open_question_text}

        {closed_context}

        Respuesta abierta:
        {open_text}

        Dimensiones a revisar:
        {'\n\n'.join(dimension_blocks)}
        """
    ).strip()


def build_schema(spec: QuestionSpec, dimensions: list[DimensionSpec]) -> dict[str, Any]:
    properties: dict[str, Any] = {
        "nota_general": {
            "type": "string",
            "description": "Nota breve opcional sobre el caso completo.",
        }
    }
    required = ["nota_general"]
    for dimension in dimensions:
        allowed_codes = [code["code_id"] for code in dimension.codes]
        properties[dimension.field] = {
            "type": "object",
            "additionalProperties": False,
            "properties": {
                "codigo_revisado": {
                    "anyOf": [
                        {"type": "string", "enum": allowed_codes},
                        {"type": "null"},
                    ]
                },
                "estado_revision": {"type": "string", "enum": REVIEW_STATES},
                "tipo_problema": {"type": "string", "enum": PROBLEM_TYPES},
                "justificacion_breve": {"type": "string"},
                "requiere_cambio_codebook": {"type": "boolean"},
            },
            "required": [
                "codigo_revisado",
                "estado_revision",
                "tipo_problema",
                "justificacion_breve",
                "requiere_cambio_codebook",
            ],
        }
        required.append(dimension.field)
    return {
        "name": f"auditoria_{spec.question}",
        "strict": True,
        "schema": {
            "type": "object",
            "additionalProperties": False,
            "properties": properties,
            "required": required,
        },
    }


def normalize_review_result(
    result: dict[str, Any],
    current_code: str | None,
    allowed_codes: set[str],
    code_id_to_label: dict[str, str],
) -> dict[str, Any]:
    state = result.get("estado_revision")
    problem = result.get("tipo_problema")
    reviewed = result.get("codigo_revisado")
    reviewed = clean_code_value(reviewed) if reviewed is not None else None
    if reviewed is not None and reviewed in code_id_to_label:
        reviewed = code_id_to_label[reviewed]
    justification = normalize_space(result.get("justificacion_breve"))
    change_codebook = bool(result.get("requiere_cambio_codebook"))

    if state not in REVIEW_STATES:
        raise RuntimeError(f"Estado de revisión inválido: {state}")
    if problem not in PROBLEM_TYPES:
        raise RuntimeError(f"Tipo de problema inválido: {problem}")
    if reviewed is not None and reviewed not in allowed_codes:
        raise RuntimeError(f"Código revisado inválido: {reviewed}")

    if state == "MISSING_VALIDO":
        reviewed = None
        change_codebook = False
        if problem == "NINGUNO":
            problem = "NINGUNO"
        if not justification:
            justification = "Respuesta vacía o sin contenido interpretable."
    elif state == "SIN_ENCAJE_CODEBOOK":
        reviewed = None
        change_codebook = True
        if problem == "NINGUNO":
            problem = "NINGUNO"
        if not justification:
            justification = "Respuesta sustantiva sin encaje limpio en el codebook."
    elif state == "MANTENER":
        if current_code is None:
            state = "CORREGIR"
            if reviewed is None:
                problem = "ERROR_PIPELINE"
                justification = justification or "El código actual estaba vacío sin que la respuesta fuera vacía."
            else:
                problem = "ERROR_PIPELINE"
        reviewed = current_code if current_code in allowed_codes else reviewed
        if reviewed is None:
            raise RuntimeError("MANTENER requiere un código actual válido.")
        if problem == "NINGUNO":
            problem = "NINGUNO"
        change_codebook = False
    elif state == "CORREGIR":
        if reviewed == current_code and current_code is not None:
            state = "MANTENER"
            problem = "NINGUNO"
        elif current_code is None and reviewed is not None and problem == "NINGUNO":
            problem = "SUBCODIFICACION"
        if not justification:
            justification = "Se ajusta mejor otro código o corresponde retirar una sobrecodificación."
        if reviewed is None and problem == "NINGUNO":
            problem = "SOBRECODIFICACION"

    return {
        "codigo_revisado": reviewed,
        "estado_revision": state,
        "tipo_problema": problem,
        "justificacion_breve": justification,
        "requiere_cambio_codebook": change_codebook,
    }


def review_case_with_openai(
    client: OpenAI,
    row: dict[str, str],
    spec: QuestionSpec,
    dimensions: list[DimensionSpec],
    current_codes: dict[str, str | None],
    systemic_alerts: dict[tuple[str, str], dict[str, Any]],
    model: str,
    reasoning_effort: str,
    verbosity: str,
) -> dict[str, Any]:
    prompt = build_prompt_payload(row, spec, dimensions, current_codes, systemic_alerts)
    schema = build_schema(spec, dimensions)

    response = client.chat.completions.create(
        model=model,
        reasoning_effort=reasoning_effort,
        verbosity=verbosity,
        max_completion_tokens=1200,
        response_format={
            "type": "json_schema",
            "json_schema": schema,
        },
        messages=[
            {
                "role": "system",
                "content": (
                    "Eres un auditor metodológico de codificación cualitativa. "
                    "Devuelve exclusivamente JSON válido según el esquema."
                ),
            },
            {"role": "user", "content": prompt},
        ],
    )
    content = response.choices[0].message.content
    if not content:
        raise RuntimeError("La API no devolvió contenido.")
    return json.loads(content)


def review_case_with_retries(
    client: OpenAI,
    row: dict[str, str],
    spec: QuestionSpec,
    dimensions: list[DimensionSpec],
    current_codes: dict[str, str | None],
    systemic_alerts: dict[tuple[str, str], dict[str, Any]],
    model: str,
    reasoning_effort: str,
    verbosity: str,
    max_retries: int,
) -> dict[str, Any]:
    for attempt in range(1, max_retries + 1):
        try:
            return review_case_with_openai(
                client=client,
                row=row,
                spec=spec,
                dimensions=dimensions,
                current_codes=current_codes,
                systemic_alerts=systemic_alerts,
                model=model,
                reasoning_effort=reasoning_effort,
                verbosity=verbosity,
            )
        except Exception as exc:
            if attempt == max_retries:
                raise
            sleep_for = min(2 ** attempt, 20)
            print(
                f"[retry] {spec.question} {row['numero']} intento {attempt}/{max_retries} falló: {exc}. "
                f"Esperando {sleep_for}s.",
                file=sys.stderr,
            )
            time.sleep(sleep_for)
    raise RuntimeError("Se agotaron los reintentos.")


def auto_missing_result() -> dict[str, Any]:
    return {
        "codigo_revisado": None,
        "estado_revision": "MISSING_VALIDO",
        "tipo_problema": "NINGUNO",
        "justificacion_breve": "Respuesta vacía o sin contenido interpretable.",
        "requiere_cambio_codebook": False,
    }


def question_key(spec: QuestionSpec, numero: str) -> str:
    return f"{spec.question}|{numero}"


def build_review_payload(
    rows: list[dict[str, str]],
    specs: list[QuestionSpec],
    current_outputs: dict[tuple[str, str], dict[str, str | None]],
    systemic_alerts: dict[tuple[str, str], dict[str, Any]],
    mode: str,
    model: str,
    reasoning_effort: str,
    verbosity: str,
    max_retries: int,
    limit: int | None,
    only_questions: set[str] | None,
    force_refresh: bool,
) -> tuple[list[dict[str, Any]], dict[str, dict[str, Any]]]:
    cached = load_cache()
    client = None
    if mode == "review":
        api_key = os.getenv("OPENAI_API_KEY")
        if not api_key:
            raise RuntimeError("No se encontró OPENAI_API_KEY. Usa --mode prepare o exporta la clave.")
        client = OpenAI(api_key=api_key, max_retries=0)

    reviewed_cases = 0
    review_lookup: dict[str, dict[str, Any]] = {}

    for spec in specs:
        if only_questions and spec.question not in only_questions:
            continue
        dimensions = load_codebook_dimensions(spec.codebook_file, spec.category_to_field)
        for row in rows:
            if limit is not None and reviewed_cases >= limit:
                break
            key = question_key(spec, row["numero"])
            open_text = row.get(spec.open_col)

            if text_is_missing(open_text):
                result = {"nota_general": "Respuesta vacía o no interpretable."}
                for dimension in dimensions:
                    result[dimension.field] = auto_missing_result()
                review_lookup[key] = result
                reviewed_cases += 1
                continue

            if not force_refresh and key in cached:
                review_lookup[key] = cached[key]["result"]
                reviewed_cases += 1
                continue

            if mode != "review":
                pending = {"nota_general": "Pendiente de revisión semántica."}
                for dimension in dimensions:
                    pending[dimension.field] = {
                        "codigo_revisado": None,
                        "estado_revision": "",
                        "tipo_problema": "",
                        "justificacion_breve": "",
                        "requiere_cambio_codebook": False,
                    }
                review_lookup[key] = pending
                reviewed_cases += 1
                continue

            current_codes = current_outputs[(spec.question, row["numero"])]
            result = review_case_with_retries(
                client=client,
                row=row,
                spec=spec,
                dimensions=dimensions,
                current_codes=current_codes,
                systemic_alerts=systemic_alerts,
                model=model,
                reasoning_effort=reasoning_effort,
                verbosity=verbosity,
                max_retries=max_retries,
            )
            review_lookup[key] = result
            append_cache(
                {
                    "key": key,
                    "numero": row["numero"],
                    "question": spec.question,
                    "timestamp": CURRENT_TIMESTAMP,
                    "model": model,
                    "result": result,
                }
            )
            reviewed_cases += 1

            if reviewed_cases % 10 == 0:
                print(f"[review] casos procesados: {reviewed_cases}", file=sys.stderr)

        if limit is not None and reviewed_cases >= limit:
            break

    return rows, review_lookup


def build_long_rows(
    rows: list[dict[str, str]],
    specs: list[QuestionSpec],
    current_outputs: dict[tuple[str, str], dict[str, str | None]],
    systemic_alerts: dict[tuple[str, str], dict[str, Any]],
    review_lookup: dict[str, dict[str, Any]],
) -> list[dict[str, Any]]:
    long_rows: list[dict[str, Any]] = []
    for spec in specs:
        dimensions = load_codebook_dimensions(spec.codebook_file, spec.category_to_field)
        for row in rows:
            key = question_key(spec, row["numero"])
            review = review_lookup.get(key, {"nota_general": ""})
            current_codes = current_outputs[(spec.question, row["numero"])]
            closed_text = row.get(f"{spec.closed_col}_texto") if spec.closed_col else None

            for dimension in dimensions:
                current_code = current_codes.get(dimension.field)
                raw_result = review.get(dimension.field)
                if not raw_result:
                    normalized = {
                        "codigo_revisado": None,
                        "estado_revision": "",
                        "tipo_problema": "",
                        "justificacion_breve": "",
                        "requiere_cambio_codebook": False,
                    }
                else:
                    if raw_result.get("estado_revision") == "":
                        normalized = {
                            "codigo_revisado": None,
                            "estado_revision": "",
                            "tipo_problema": "",
                            "justificacion_breve": "",
                            "requiere_cambio_codebook": False,
                        }
                    else:
                        normalized = normalize_review_result(
                            result=raw_result,
                            current_code=current_code,
                            allowed_codes={code["codigo"] for code in dimension.codes},
                            code_id_to_label={
                                code["code_id"]: code["codigo"] for code in dimension.codes
                            },
                        )

                long_rows.append(
                    {
                        **{col: row.get(col) for col in BASIC_METADATA},
                        "pregunta": spec.question,
                        "tema": spec.tema,
                        "pregunta_texto": spec.open_question_text,
                        "dimension": dimension.category,
                        "campo_salida": dimension.field,
                        "respuesta_abierta": row.get(spec.open_col),
                        "respuesta_cerrada_texto": closed_text,
                        "codigo_actual": current_code,
                        "codigo_revisado": normalized["codigo_revisado"],
                        "estado_revision": normalized["estado_revision"],
                        "tipo_problema": normalized["tipo_problema"],
                        "justificacion_breve": normalized["justificacion_breve"],
                        "requiere_cambio_codebook": normalized["requiere_cambio_codebook"],
                        "alerta_sistemica": systemic_alerts[(spec.question, dimension.field)]["alert_text"],
                        "nota_general_caso": review.get("nota_general", ""),
                    }
                )
    return long_rows


def build_summary_rows(long_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in long_rows:
        grouped[(row["pregunta"], row["dimension"])].append(row)

    summary_rows: list[dict[str, Any]] = []
    for (question, dimension), items in sorted(grouped.items()):
        state_counts = Counter(item["estado_revision"] or "PENDIENTE" for item in items)
        type_counts = Counter(item["tipo_problema"] or "PENDIENTE" for item in items)
        changes = sum(1 for item in items if item["estado_revision"] == "CORREGIR")
        required_codebook = sum(1 for item in items if item["requiere_cambio_codebook"])
        current_counter = Counter(item["codigo_actual"] or "NA" for item in items)
        reviewed_counter = Counter(item["codigo_revisado"] or "NA" for item in items)
        current_top = current_counter.most_common(1)[0][0]
        reviewed_top = reviewed_counter.most_common(1)[0][0]

        summary_rows.append(
            {
                "pregunta": question,
                "dimension": dimension,
                "n_casos": len(items),
                "mantener": state_counts.get("MANTENER", 0),
                "corregir": state_counts.get("CORREGIR", 0),
                "missing_valido": state_counts.get("MISSING_VALIDO", 0),
                "sin_encaje_codebook": state_counts.get("SIN_ENCAJE_CODEBOOK", 0),
                "pendiente": state_counts.get("PENDIENTE", 0),
                "tasa_cambio": round(changes / len(items), 4),
                "requiere_cambio_codebook": required_codebook,
                "subcodificacion": type_counts.get("SUBCODIFICACION", 0),
                "sobrecodificacion": type_counts.get("SOBRECODIFICACION", 0),
                "sesgo_de_default": type_counts.get("SESGO_DE_DEFAULT", 0),
                "error_pipeline": type_counts.get("ERROR_PIPELINE", 0),
                "codigo_actual_top": current_top,
                "codigo_revisado_top": reviewed_top,
            }
        )
    return summary_rows


def build_systemic_alert_rows(
    long_rows: list[dict[str, Any]],
    systemic_alerts: dict[tuple[str, str], dict[str, Any]],
) -> list[dict[str, Any]]:
    summary_by_dim: dict[tuple[str, str], Counter] = defaultdict(Counter)
    for row in long_rows:
        summary_by_dim[(row["pregunta"], row["campo_salida"])][row["estado_revision"] or "PENDIENTE"] += 1
        summary_by_dim[(row["pregunta"], row["campo_salida"])][f"tipo::{row['tipo_problema'] or 'PENDIENTE'}"] += 1

    rows: list[dict[str, Any]] = []
    for key, alert in sorted(systemic_alerts.items()):
        if not alert["alert_text"] and key not in SYSTEMIC_DIMENSIONS_OF_INTEREST:
            continue
        counter = summary_by_dim[key]
        rows.append(
            {
                "pregunta": alert["question"],
                "campo_salida": alert["field"],
                "dimension": alert["dimension"],
                "alerta_inicial": alert["alert_text"] or "Sin alerta sistémica detectada.",
                "mantener": counter.get("MANTENER", 0),
                "corregir": counter.get("CORREGIR", 0),
                "missing_valido": counter.get("MISSING_VALIDO", 0),
                "sin_encaje_codebook": counter.get("SIN_ENCAJE_CODEBOOK", 0),
                "subcodificacion": counter.get("tipo::SUBCODIFICACION", 0),
                "sesgo_de_default": counter.get("tipo::SESGO_DE_DEFAULT", 0),
                "error_pipeline": counter.get("tipo::ERROR_PIPELINE", 0),
            }
        )
    return rows


def build_codebook_alert_rows(long_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        row
        for row in long_rows
        if row["estado_revision"] == "SIN_ENCAJE_CODEBOOK" or row["requiere_cambio_codebook"]
    ]


def build_coherence_rows(
    rows: list[dict[str, str]],
    long_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    review_map: dict[tuple[str, str], dict[str, Any]] = {}
    for row in long_rows:
        review_map[(row["numero"], row["campo_salida"])] = row

    coherence_rows: list[dict[str, Any]] = []
    for row in rows:
        numero = row["numero"]
        pairs = [
            (
                "q2-q4",
                review_map.get((numero, "postura_general_q2")),
                review_map.get((numero, "postura_debate_q4")),
                "postura",
            ),
            (
                "q5-q6",
                row.get("q5_texto"),
                review_map.get((numero, "evaluacion_q6")),
                {
                    "Me resultaron creíbles": "Creíbles",
                    "No me resultaron creíbles": "No creíbles",
                    "No sabría decir": "No sabe / Dudoso",
                },
            ),
            (
                "q3-q4",
                row.get("q3_texto"),
                review_map.get((numero, "evaluacion_oposicion_q4")),
                {
                    "Está actuando de manera correcta": "Correcta",
                    "No está actuando de manera correcta": "Incorrecta",
                    "No tengo una opinión formada": "Sin Definir",
                },
            ),
            (
                "q9-q10",
                row.get("q9_texto"),
                review_map.get((numero, "codigo_q10")),
                "rehabilitacion",
            ),
            (
                "q11-q12",
                row.get("q11_texto"),
                review_map.get((numero, "codigo_q12")),
                "trabajo_estudio",
            ),
        ]

        for pair_name, left, right, rule in pairs:
            if rule == "postura":
                q2_review = left
                q4_review = right
                if not q2_review or not q4_review:
                    continue
                q2_code = q2_review["codigo_revisado"]
                q4_code = q4_review["codigo_revisado"]
                consistent = {
                    ("Pro-Gobierno / Ministro", "Pro-Ministro"),
                    ("Pro-Bordaberry / Oposición", "Pro-Bordaberry"),
                    ("Rechazo / Punitivismo extremo", "Pro-Bordaberry"),
                    ("Rechazo / Punitivismo extremo", "Punitivismo Radical"),
                    ("Mixta / Punto Medio", "Mixta / Gris"),
                    ("NS/NC / Desinformado", "NS / NC"),
                }
                if q2_code and q4_code and (q2_code, q4_code) not in consistent:
                    coherence_rows.append(
                        {
                            "numero": numero,
                            "par": pair_name,
                            "respuesta_cerrada": "",
                            "campo_salida": f"{q2_review['campo_salida']} / {q4_review['campo_salida']}",
                            "codigo_revisado": f"{q2_code} / {q4_code}",
                            "respuesta_abierta": f"q2: {q2_review['respuesta_abierta']} || q4: {q4_review['respuesta_abierta']}",
                            "justificacion_breve": "La postura revisada en q2 y q4 no queda alineada entre sí.",
                        }
                    )
                continue

            closed_text = left
            reviewed_dimension = right
            if not closed_text or not reviewed_dimension:
                continue
            revised_code = reviewed_dimension["codigo_revisado"]

            if isinstance(rule, dict):
                expected = rule.get(closed_text)
                if expected and revised_code and revised_code != expected:
                    coherence_rows.append(
                        {
                            "numero": numero,
                            "par": pair_name,
                            "respuesta_cerrada": closed_text,
                            "campo_salida": reviewed_dimension["campo_salida"],
                            "codigo_revisado": revised_code,
                            "respuesta_abierta": reviewed_dimension["respuesta_abierta"],
                            "justificacion_breve": reviewed_dimension["justificacion_breve"],
                        }
                    )
                continue

            if rule == "rehabilitacion":
                positive_codes = {
                    "Reducción de Reincidencia",
                    "Derecho y Humanidad",
                    "Función Institucional (Deber Ser)",
                    "Reinserción Laboral/Educativa",
                    "Escepticismo / Voluntad Individual",
                    "Crítica al Sistema Actual",
                    "Costo-Beneficio Social",
                }
                negative_codes = {"Punitivismo / Inutilidad"}
                closed_positive = {"Muy importante", "Algo importante"}
                closed_negative = {"Poco importante", "Nada importante"}
                if closed_text in closed_positive and revised_code in negative_codes:
                    coherence_rows.append(
                        {
                            "numero": numero,
                            "par": pair_name,
                            "respuesta_cerrada": closed_text,
                            "campo_salida": reviewed_dimension["campo_salida"],
                            "codigo_revisado": revised_code,
                            "respuesta_abierta": reviewed_dimension["respuesta_abierta"],
                            "justificacion_breve": "La cerrada es pro-rehabilitación y la abierta quedó en código abiertamente anti-rehabilitación.",
                        }
                    )
                elif closed_text in closed_negative and revised_code in positive_codes:
                    coherence_rows.append(
                        {
                            "numero": numero,
                            "par": pair_name,
                            "respuesta_cerrada": closed_text,
                            "campo_salida": reviewed_dimension["campo_salida"],
                            "codigo_revisado": revised_code,
                            "respuesta_abierta": reviewed_dimension["respuesta_abierta"],
                            "justificacion_breve": "La cerrada es anti-rehabilitación y la abierta quedó en un racional claramente pro-rehabilitación.",
                        }
                    )
                continue

            if rule == "trabajo_estudio":
                positive_codes = {
                    "Combate al Ocio",
                    "Dignificación y Autoestima",
                    "Herramientas de Reinserción",
                    "Hábitos y Disciplina",
                    "Justicia Social / Brecha Educativa",
                    "Meritocracia y Selección",
                }
                negative_codes = {
                    "Injusticia Percibida (Exclusión)",
                    "Escepticismo sobre el Cambio",
                }
                closed_positive = {"Muy importante", "Algo importante"}
                closed_negative = {"Poco importante", "Nada importante"}
                if closed_text in closed_positive and revised_code in negative_codes:
                    coherence_rows.append(
                        {
                            "numero": numero,
                            "par": pair_name,
                            "respuesta_cerrada": closed_text,
                            "campo_salida": reviewed_dimension["campo_salida"],
                            "codigo_revisado": revised_code,
                            "respuesta_abierta": reviewed_dimension["respuesta_abierta"],
                            "justificacion_breve": "La cerrada apoya trabajo/estudio y la abierta quedó en un código crítico o escéptico.",
                        }
                    )
                elif closed_text in closed_negative and revised_code in positive_codes:
                    coherence_rows.append(
                        {
                            "numero": numero,
                            "par": pair_name,
                            "respuesta_cerrada": closed_text,
                            "campo_salida": reviewed_dimension["campo_salida"],
                            "codigo_revisado": revised_code,
                            "respuesta_abierta": reviewed_dimension["respuesta_abierta"],
                            "justificacion_breve": "La cerrada rechaza la importancia y la abierta quedó en un racional claramente favorable.",
                        }
                    )
    return coherence_rows


def append_sheet(ws, rows: list[dict[str, Any]], headers: list[str]) -> None:
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for index, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=index)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
    for row in rows:
        ws.append([row.get(header) for header in headers])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for column_cells in ws.columns:
        max_len = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells[: min(len(column_cells), 60)]:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[column_letter].width = min(max(max_len + 2, 12), 60)


def build_reviewed_wide_rows(
    rows: list[dict[str, str]],
    specs: list[QuestionSpec],
    long_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    long_lookup: dict[tuple[str, str], dict[str, Any]] = {}
    for row in long_rows:
        long_lookup[(row["numero"], row["campo_salida"])] = row

    reviewed_rows: list[dict[str, Any]] = []
    for row in rows:
        base = dict(row)
        for question in OPTION_MAPS:
            if question in base:
                base[question] = base.get(f"{question}_texto")
        reviewed_rows.append(base)

    for reviewed_row in reviewed_rows:
        numero = reviewed_row["numero"]
        for spec in specs:
            dimensions = load_codebook_dimensions(spec.codebook_file, spec.category_to_field)
            for dimension in dimensions:
                lookup = long_lookup[(numero, dimension.field)]
                reviewed_row[f"{dimension.field}_actual"] = lookup["codigo_actual"]
                reviewed_row[f"{dimension.field}_revisado"] = lookup["codigo_revisado"]
                reviewed_row[f"{dimension.field}_estado_revision"] = lookup["estado_revision"]
                reviewed_row[f"{dimension.field}_tipo_problema"] = lookup["tipo_problema"]

    ordered_columns = BASIC_METADATA.copy()
    for question_number in range(1, 13):
        q = f"q{question_number}"
        ordered_columns.append(q)
        for spec in specs:
            if spec.question != q:
                continue
            dimensions = load_codebook_dimensions(spec.codebook_file, spec.category_to_field)
            for dimension in dimensions:
                ordered_columns.extend(
                    [
                        f"{dimension.field}_actual",
                        f"{dimension.field}_revisado",
                        f"{dimension.field}_estado_revision",
                        f"{dimension.field}_tipo_problema",
                    ]
                )

    final_rows: list[dict[str, Any]] = []
    for row in reviewed_rows:
        final_rows.append({column: row.get(column) for column in ordered_columns})
    return final_rows


def write_workbooks(
    long_rows: list[dict[str, Any]],
    summary_rows: list[dict[str, Any]],
    systemic_rows: list[dict[str, Any]],
    codebook_rows: list[dict[str, Any]],
    coherence_rows: list[dict[str, Any]],
    reviewed_wide_rows: list[dict[str, Any]],
) -> None:
    audit_wb = Workbook()
    audit_wb.remove(audit_wb.active)

    ws_summary = audit_wb.create_sheet("resumen")
    append_sheet(ws_summary, summary_rows, list(summary_rows[0].keys()) if summary_rows else ["pregunta"])

    ws_systemic = audit_wb.create_sheet("alertas_sistemicas")
    append_sheet(
        ws_systemic,
        systemic_rows,
        list(systemic_rows[0].keys()) if systemic_rows else ["pregunta"],
    )

    ws_long = audit_wb.create_sheet("auditoria_long")
    append_sheet(ws_long, long_rows, list(long_rows[0].keys()) if long_rows else ["numero"])

    ws_codebook = audit_wb.create_sheet("alertas_codebook")
    append_sheet(
        ws_codebook,
        codebook_rows,
        list(codebook_rows[0].keys()) if codebook_rows else list(long_rows[0].keys()),
    )

    ws_coherence = audit_wb.create_sheet("coherencia_pares")
    append_sheet(
        ws_coherence,
        coherence_rows,
        list(coherence_rows[0].keys()) if coherence_rows else ["numero", "par"],
    )

    audit_wb.save(AUDIT_WORKBOOK)

    reviewed_wb = Workbook()
    ws = reviewed_wb.active
    ws.title = "R8_codificada_revisada"
    append_sheet(
        ws,
        reviewed_wide_rows,
        list(reviewed_wide_rows[0].keys()) if reviewed_wide_rows else ["numero"],
    )
    reviewed_wb.save(REVIEWED_WIDE_WORKBOOK)


def build_memo(
    long_rows: list[dict[str, Any]],
    summary_rows: list[dict[str, Any]],
    systemic_rows: list[dict[str, Any]],
    coherence_rows: list[dict[str, Any]],
) -> str:
    total_decisions = len(long_rows)
    state_counts = Counter(row["estado_revision"] or "PENDIENTE" for row in long_rows)
    type_counts = Counter(row["tipo_problema"] or "PENDIENTE" for row in long_rows)
    change_dims = sorted(summary_rows, key=lambda row: row["tasa_cambio"], reverse=True)[:8]
    codebook_cases = [row for row in long_rows if row["requiere_cambio_codebook"]]

    systemic_lines = []
    for row in systemic_rows:
        if row["alerta_inicial"] == "Sin alerta sistémica detectada." and (row["pregunta"], row["campo_salida"]) not in SYSTEMIC_DIMENSIONS_OF_INTEREST:
            continue
        systemic_lines.append(
            f"- `{row['pregunta']}/{row['campo_salida']}`: {row['alerta_inicial']} "
            f"Tras la revisión: corregir={row['corregir']}, subcodificación={row['subcodificacion']}, "
            f"sesgo_default={row['sesgo_de_default']}, error_pipeline={row['error_pipeline']}."
        )

    change_lines = [
        f"- `{row['pregunta']} / {row['dimension']}`: tasa_cambio={row['tasa_cambio']:.1%}, "
        f"actual_top=`{row['codigo_actual_top']}`, revisado_top=`{row['codigo_revisado_top']}`."
        for row in change_dims
    ]

    memo = [
        "# Revisión integral R8 2026",
        "",
        f"Fecha de generación: {CURRENT_DATE}",
        "",
        "## Cobertura",
        f"- Decisiones auditadas: {total_decisions}",
        f"- `MANTENER`: {state_counts.get('MANTENER', 0)}",
        f"- `CORREGIR`: {state_counts.get('CORREGIR', 0)}",
        f"- `MISSING_VALIDO`: {state_counts.get('MISSING_VALIDO', 0)}",
        f"- `SIN_ENCAJE_CODEBOOK`: {state_counts.get('SIN_ENCAJE_CODEBOOK', 0)}",
        f"- `PENDIENTE`: {state_counts.get('PENDIENTE', 0)}",
        "",
        "## Hallazgos de pipeline y prompt",
        *(
            systemic_lines
            if systemic_lines
            else ["- No se detectaron alertas sistémicas fuertes adicionales a las ya conocidas."]
        ),
        "",
        "## Hallazgos semánticos",
        f"- `SUBCODIFICACION`: {type_counts.get('SUBCODIFICACION', 0)}",
        f"- `SOBRECODIFICACION`: {type_counts.get('SOBRECODIFICACION', 0)}",
        f"- `SESGO_DE_DEFAULT`: {type_counts.get('SESGO_DE_DEFAULT', 0)}",
        f"- `ERROR_PIPELINE`: {type_counts.get('ERROR_PIPELINE', 0)}",
        "",
        "Dimensiones con mayor tasa de cambio:",
        *(change_lines if change_lines else ["- Sin cambios registrados."]),
        "",
        "## Codebook",
        f"- Casos con `requiere_cambio_codebook = TRUE`: {len(codebook_cases)}",
        (
            "- Hay respuestas sustantivas sin encaje limpio en el codebook y conviene revisar esas categorías."
            if codebook_cases
            else "- No se detectaron casos con brecha explícita de codebook en esta corrida."
        ),
        "",
        "## Coherencia cruzada",
        f"- Alertas simples de coherencia detectadas: {len(coherence_rows)}",
        "- Estas alertas no implican error automático; sirven para un segundo control comparando cerrada y abierta.",
    ]
    return "\n".join(memo) + "\n"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Revisión integral de codificación R8.")
    parser.add_argument(
        "--mode",
        choices=["prepare", "review"],
        default="prepare",
        help="prepare arma la auditoría sin adjudicación semántica; review usa OpenAI para revisar caso por caso.",
    )
    parser.add_argument("--model", default=MODEL_DEFAULT)
    parser.add_argument("--reasoning-effort", default="medium")
    parser.add_argument("--verbosity", default="low")
    parser.add_argument("--max-retries", type=int, default=4)
    parser.add_argument("--limit", type=int, default=None)
    parser.add_argument(
        "--questions",
        default=None,
        help="Lista separada por comas, por ejemplo q2,q4,q6.",
    )
    parser.add_argument(
        "--force-refresh",
        action="store_true",
        help="Ignora cache existente y vuelve a revisar los casos.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    specs = question_specs()
    rows = load_csv(TRANSCRIPTION_FILE)
    enrich_closed_answers(rows)
    current_outputs = load_current_outputs(specs)
    systemic_alerts = compute_systemic_alerts(rows, specs, current_outputs)

    only_questions = None
    if args.questions:
        only_questions = {question.strip() for question in args.questions.split(",") if question.strip()}

    rows, review_lookup = build_review_payload(
        rows=rows,
        specs=specs,
        current_outputs=current_outputs,
        systemic_alerts=systemic_alerts,
        mode=args.mode,
        model=args.model,
        reasoning_effort=args.reasoning_effort,
        verbosity=args.verbosity,
        max_retries=args.max_retries,
        limit=args.limit,
        only_questions=only_questions,
        force_refresh=args.force_refresh,
    )

    long_rows = build_long_rows(
        rows=rows,
        specs=specs,
        current_outputs=current_outputs,
        systemic_alerts=systemic_alerts,
        review_lookup=review_lookup,
    )
    summary_rows = build_summary_rows(long_rows)
    systemic_rows = build_systemic_alert_rows(long_rows, systemic_alerts)
    codebook_rows = build_codebook_alert_rows(long_rows)
    coherence_rows = build_coherence_rows(rows, long_rows)
    reviewed_wide_rows = build_reviewed_wide_rows(rows, specs, long_rows)

    write_csv(AUDIT_LONG_CSV, long_rows, list(long_rows[0].keys()))
    write_workbooks(
        long_rows=long_rows,
        summary_rows=summary_rows,
        systemic_rows=systemic_rows,
        codebook_rows=codebook_rows,
        coherence_rows=coherence_rows,
        reviewed_wide_rows=reviewed_wide_rows,
    )

    memo = build_memo(
        long_rows=long_rows,
        summary_rows=summary_rows,
        systemic_rows=systemic_rows,
        coherence_rows=coherence_rows,
    )
    REVIEW_MEMO.write_text(memo, encoding="utf-8")

    print(f"Workbook auditoría: {AUDIT_WORKBOOK}")
    print(f"CSV auditoría larga: {AUDIT_LONG_CSV}")
    print(f"Base revisada: {REVIEWED_WIDE_WORKBOOK}")
    print(f"Memo: {REVIEW_MEMO}")


if __name__ == "__main__":
    main()
