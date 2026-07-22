from __future__ import annotations

import csv
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


REPO_ROOT = Path(__file__).resolve().parents[4]
ANALYSIS_DIR = REPO_ROOT / "data" / "processed" / "analysis" / "2026" / "R8"
TRANSCRIPTION_FILE = (
    REPO_ROOT
    / "data"
    / "processed"
    / "transcriptions"
    / "output"
    / "2026"
    / "transcripcion_R8.csv"
)
AUDIT_LONG_INPUT = ANALYSIS_DIR / "R8_auditoria_long.csv"

AUDIT_LONG_OUTPUT = ANALYSIS_DIR / "R8_auditoria_long_manual.csv"
AUDIT_WORKBOOK_OUTPUT = ANALYSIS_DIR / "R8_auditoria_manual.xlsx"
FINAL_WIDE_OUTPUT = ANALYSIS_DIR / "R8_codificada_final.xlsx"
FINAL_CLEAN_OUTPUT = ANALYSIS_DIR / "R8_codificada.xlsx"
FINAL_CLEAN_CSV_OUTPUT = ANALYSIS_DIR / "R8_codificada.csv"
MANUAL_MEMO_OUTPUT = ANALYSIS_DIR / "R8_hallazgos_manual_final.md"


OPTION_MAPS = {
    "q3": {
        "A": "Está actuando de manera correcta",
        "B": "No está actuando de manera correcta",
        "C": "No tengo una opinión formada",
    },
    "q5": {
        "A": "Me resultaron creíbles",
        "B": "No me resultaron creíbles",
        "C": "No sabría decir",
    },
    "q7": {
        "A": "en la construcción de nuevas cárceles",
        "B": "en la mejora de las cárceles ya existentes",
        "C": "tanto en la construcción de cárceles como en la mejora de las ya existentes",
        "D": "NO es prioritario que el Estado uruguayo invierta en cárceles.",
    },
    "q9": {
        "A": "Muy importante",
        "B": "Algo importante",
        "C": "Ni una cosa ni la otra",
        "D": "Poco importante",
        "E": "Nada importante",
        "F": "No tengo opinión",
    },
    "q11": {
        "A": "Muy importante",
        "B": "Algo importante",
        "C": "Ni una cosa ni la otra",
        "D": "Poco importante",
        "E": "Nada importante",
        "F": "No tengo opinión",
    },
}

QUESTION_FIELD_ORDER = {
    "q1": ["nivel_exposicion_q1", "dimension_tematica_q1", "tono_q1"],
    "q2": [
        "postura_general_q2",
        "percepcion_sistema_q2",
        "argumentos_temores_q2",
        "propuestas_solucion_q2",
    ],
    "q4": [
        "postura_debate_q4",
        "evaluacion_oposicion_q4",
        "argumentos_justificacion_q4",
        "percepcion_sistema_q4",
    ],
    "q6": [
        "evaluacion_q6",
        "argumentos_desconfianza_q6",
        "argumentos_confianza_q6",
        "influencia_externa_q6",
        "contexto_otros_q6",
    ],
    "q8": [
        "prioridad_inversion_q8",
        "justificacion_rehabilitacion_q8",
        "justificacion_punitivismo_q8",
        "gestion_recursos_q8",
    ],
    "q10": ["codigo_q10"],
    "q12": ["codigo_q12"],
}

ABSENCE_FALLBACK_CODE = {
    "postura_general_q2": "NS/NC / Desinformado",
    "postura_debate_q4": "NS / NC",
    "evaluacion_oposicion_q4": "Sin Definir",
}

ABSENCE_KEEP_CODES = {
    "nivel_exposicion_q1": {"No vio ni sabía", "No vio pero sabía"},
    "tono_q1": {"Neutral / Descriptivo"},
    "postura_general_q2": {"NS/NC / Desinformado"},
    "postura_debate_q4": {"NS / NC"},
    "evaluacion_oposicion_q4": {"Sin Definir"},
    "evaluacion_q6": {"No sabe / Dudoso"},
}

ABSENCE_PATTERN = re.compile(
    r"^(No (hay|aparece|menciona|expresa|desarrolla|aporta|describe|evalúa|formula|propone|recuerda|distingue|ofrece)|"
    r"No se menciona|Solo dice|Solo reporta ausencia|Respuesta vacía)",
    re.IGNORECASE,
)
GAP_PATTERN = re.compile(
    r"(no encaja|fuera de las categor|no activa de forma específica|menciona .* pero no encaja|habla de .* pero no encaja)",
    re.IGNORECASE,
)

STATUS_MAP = {
    "MANTENER": "MANTENER_FINAL",
    "CORREGIR": "CORREGIR_FINAL",
    "MISSING_VALIDO": "MISSING_VALIDO_FINAL",
    "SIN_ENCAJE_CODEBOOK": "SIN_ENCAJE_CODEBOOK_FINAL",
}

EXPLICIT_OVERRIDES: dict[tuple[str, str], dict[str, Any]] = {
    (
        "59894583422",
        "argumentos_justificacion_q4",
    ): {
        "codigo_manual_final": None,
        "estado_manual_final": "MISSING_VALIDO_FINAL",
        "justificacion_manual_final": (
            "Corrección manual: la respuesta solo indica a quién escuchó y no justifica la "
            "actuación de la oposición en esta dimensión."
        ),
        "requiere_cambio_codebook_manual": False,
        "fuente_manual_final": "override_explicito",
    },
}


def normalize_space(value: str | None) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\r", " ").replace("\n", " ").split())


def clean_code(value: str | None) -> str | None:
    text = normalize_space(value)
    if not text:
        return None
    if text.upper() in {"NA", "N/A", "NULL"}:
        return None
    return text


def text_is_empty(value: str | None) -> bool:
    text = normalize_space(value)
    return text.lower() in {"", "na", "n/a", "...", ".", "..", "-", "--"}


def bool_from_str(value: Any) -> bool:
    return str(value).strip().lower() == "true"


def load_csv(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        return list(csv.DictReader(handle))


def write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def recode_option_letter(question: str, value: str | None) -> str | None:
    cleaned = clean_code(value)
    if cleaned is None:
        return None
    mapping = OPTION_MAPS.get(question)
    if mapping is None:
        return cleaned
    return mapping.get(cleaned, cleaned)


def append_sheet(ws, rows: list[dict[str, Any]], headers: list[str]) -> None:
    ws.append(headers)
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx)
        cell.fill = fill
        cell.font = font
    for row in rows:
        ws.append([row.get(header) for header in headers])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for column in ws.columns:
        column_letter = column[0].column_letter
        max_len = 0
        for cell in column[: min(len(column), 80)]:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[column_letter].width = min(max(max_len + 2, 12), 60)


def manualize_row(row: dict[str, str]) -> dict[str, Any]:
    field = row["campo_salida"]
    numero = row["numero"]
    reviewed_state = row["estado_revision"]
    current_code = clean_code(row.get("codigo_actual"))
    reviewed_code = clean_code(row.get("codigo_revisado"))
    justification = normalize_space(row.get("justificacion_breve"))
    requires_change = bool_from_str(row.get("requiere_cambio_codebook"))

    result = dict(row)

    if reviewed_state == "MANTENER":
        result.update(
            {
                "codigo_manual_final": reviewed_code or current_code,
                "estado_manual_final": "MANTENER_FINAL",
                "justificacion_manual_final": (
                    "Congelado desde la auditoría previa: el código revisado ya se considera correcto."
                ),
                "requiere_cambio_codebook_manual": requires_change,
                "fuente_manual_final": "congelado_mantener",
            }
        )
        return result

    override = EXPLICIT_OVERRIDES.get((numero, field))
    if override is not None:
        result.update(override)
        return result

    if ABSENCE_PATTERN.search(justification) and not GAP_PATTERN.search(justification):
        fallback = ABSENCE_FALLBACK_CODE.get(field)
        keep_codes = ABSENCE_KEEP_CODES.get(field, set())

        if reviewed_code in keep_codes or (fallback is not None and reviewed_code == fallback):
            result.update(
                {
                    "codigo_manual_final": reviewed_code,
                    "estado_manual_final": "MANTENER_FINAL",
                    "justificacion_manual_final": (
                        "Corrección manual validada: la respuesta no activa un contenido sustantivo adicional y "
                        "el código neutro/ausencia ya era el adecuado."
                    ),
                    "requiere_cambio_codebook_manual": False,
                    "fuente_manual_final": "validacion_codigo_neutro",
                }
            )
            return result

        if fallback is not None:
            result.update(
                {
                    "codigo_manual_final": fallback,
                    "estado_manual_final": "CORREGIR_FINAL",
                    "justificacion_manual_final": (
                        "Corrección manual: la respuesta no activa esta dimensión de forma sustantiva; "
                        "corresponde usar el código neutro/indeterminado de la propia dimensión."
                    ),
                    "requiere_cambio_codebook_manual": False,
                    "fuente_manual_final": "fallback_ausencia_dimension",
                }
            )
            return result

        result.update(
            {
                "codigo_manual_final": None,
                "estado_manual_final": "MISSING_VALIDO_FINAL",
                "justificacion_manual_final": (
                    "Corrección manual: la respuesta no activa esta dimensión; se trata de ausencia válida, "
                    "no de una brecha del codebook."
                ),
                "requiere_cambio_codebook_manual": False,
                "fuente_manual_final": "ausencia_dimension",
            }
        )
        return result

    if reviewed_state == "CORREGIR" and reviewed_code is None:
        result.update(
            {
                "codigo_manual_final": None,
                "estado_manual_final": "MISSING_VALIDO_FINAL",
                "justificacion_manual_final": (
                    "Corrección manual: la revisión previa dejó la fila sin código sustantivo y el texto no activa "
                    "esta dimensión de forma codificable."
                ),
                "requiere_cambio_codebook_manual": False,
                "fuente_manual_final": "corregir_sin_codigo",
            }
        )
        return result

    result.update(
        {
            "codigo_manual_final": reviewed_code,
            "estado_manual_final": STATUS_MAP.get(reviewed_state, reviewed_state),
            "justificacion_manual_final": (
                justification or "Se conserva la adjudicación revisada previa como correcta en el pase manual focalizado."
            ),
            "requiere_cambio_codebook_manual": requires_change,
            "fuente_manual_final": "conservar_revision_previa",
        }
    )
    return result


def build_summary(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        grouped[(row["pregunta"], row["campo_salida"])].append(row)

    summary: list[dict[str, Any]] = []
    for (question, field), items in sorted(grouped.items(), key=lambda kv: (int(kv[0][0][1:]), kv[0][1])):
        counter = Counter(item["estado_manual_final"] for item in items)
        sources = Counter(item["fuente_manual_final"] for item in items)
        changed = sum(
            1
            for item in items
            if clean_code(item.get("codigo_manual_final")) != clean_code(item.get("codigo_revisado"))
            or item.get("estado_manual_final") != f"{item.get('estado_revision')}_FINAL"
        )
        summary.append(
            {
                "pregunta": question,
                "campo_salida": field,
                "n": len(items),
                "mantener_final": counter.get("MANTENER_FINAL", 0),
                "corregir_final": counter.get("CORREGIR_FINAL", 0),
                "missing_final": counter.get("MISSING_VALIDO_FINAL", 0),
                "sin_encaje_final": counter.get("SIN_ENCAJE_CODEBOOK_FINAL", 0),
                "cambios_manual_vs_revision": changed,
                "congelado_mantener": sources.get("congelado_mantener", 0),
                "fallback_ausencia_dimension": sources.get("fallback_ausencia_dimension", 0),
                "ausencia_dimension": sources.get("ausencia_dimension", 0),
                "override_explicito": sources.get("override_explicito", 0),
                "conservar_revision_previa": sources.get("conservar_revision_previa", 0),
            }
        )
    return summary


def build_coherence(rows_long: list[dict[str, Any]], transcriptions: list[dict[str, str]]) -> list[dict[str, Any]]:
    lookup = {(row["numero"], row["campo_salida"]): row for row in rows_long}
    coherence: list[dict[str, Any]] = []

    q3_map = {
        "Está actuando de manera correcta": "Correcta",
        "No está actuando de manera correcta": "Incorrecta",
        "No tengo una opinión formada": "Sin Definir",
    }
    q5_map = {
        "Me resultaron creíbles": "Creíbles",
        "No me resultaron creíbles": "No creíbles",
        "No sabría decir": "No sabe / Dudoso",
    }

    for row in transcriptions:
        numero = row["numero"]
        q2 = lookup.get((numero, "postura_general_q2"))
        q4 = lookup.get((numero, "postura_debate_q4"))
        if q2 and q4:
            q2_code = clean_code(q2["codigo_manual_final"])
            q4_code = clean_code(q4["codigo_manual_final"])
            consistent = {
                ("Pro-Gobierno / Ministro", "Pro-Ministro"),
                ("Pro-Bordaberry / Oposición", "Pro-Bordaberry"),
                ("Rechazo / Punitivismo extremo", "Pro-Bordaberry"),
                ("Rechazo / Punitivismo extremo", "Punitivismo Radical"),
                ("Mixta / Punto Medio", "Mixta / Gris"),
                ("NS/NC / Desinformado", "NS / NC"),
            }
            if q2_code and q4_code and (q2_code, q4_code) not in consistent:
                coherence.append(
                    {
                        "numero": numero,
                        "par": "q2-q4",
                        "respuesta_cerrada": "",
                        "codigo_manual_final": f"{q2_code} / {q4_code}",
                        "respuesta_abierta": f"q2: {row.get('q2')} || q4: {row.get('q4')}",
                        "nota": "Persisten posturas no alineadas entre q2 y q4.",
                    }
                )

        q3_text = row.get("q3_texto")
        eval_q4 = lookup.get((numero, "evaluacion_oposicion_q4"))
        if q3_text and eval_q4:
            eval_code = clean_code(eval_q4["codigo_manual_final"])
            expected = q3_map.get(q3_text)
            if expected and eval_code and eval_code != expected:
                coherence.append(
                    {
                        "numero": numero,
                        "par": "q3-q4",
                        "respuesta_cerrada": q3_text,
                        "codigo_manual_final": eval_code,
                        "respuesta_abierta": row.get("q4"),
                        "nota": "La cerrada q3 y la evaluación manual de q4 no coinciden.",
                    }
                )

        q5_text = row.get("q5_texto")
        eval_q6 = lookup.get((numero, "evaluacion_q6"))
        if q5_text and eval_q6:
            eval_code = clean_code(eval_q6["codigo_manual_final"])
            expected = q5_map.get(q5_text)
            if expected and eval_code and eval_code != expected:
                coherence.append(
                    {
                        "numero": numero,
                        "par": "q5-q6",
                        "respuesta_cerrada": q5_text,
                        "codigo_manual_final": eval_code,
                        "respuesta_abierta": row.get("q6"),
                        "nota": "La cerrada q5 y la evaluación manual de q6 no coinciden.",
                    }
                )

    return coherence


def build_final_wide(
    transcriptions: list[dict[str, str]],
    rows_long: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    lookup = {(row["numero"], row["campo_salida"]): row for row in rows_long}

    base_columns = [
        "nombre",
        "edad",
        "genero",
        "numero",
        "departamento",
        "n_educativo",
        "oficio",
        "voto",
        "segmento",
        "voto2",
        "etiqueta",
    ]

    final_rows: list[dict[str, Any]] = []
    for row in transcriptions:
        out = {column: row.get(column) for column in base_columns}
        for question_number in range(1, 13):
            question = f"q{question_number}"
            value = row.get(question)
            if question in OPTION_MAPS:
                value = row.get(f"{question}_texto")
            out[question] = value
            if question in QUESTION_FIELD_ORDER:
                for field in QUESTION_FIELD_ORDER[question]:
                    audit = lookup[(row["numero"], field)]
                    out[f"{field}_manual_final"] = audit["codigo_manual_final"]
                    out[f"{field}_actual"] = audit["codigo_actual"]
                    out[f"{field}_revisado"] = audit["codigo_revisado"]
                    out[f"{field}_estado_manual_final"] = audit["estado_manual_final"]
                    out[f"{field}_fuente_manual_final"] = audit["fuente_manual_final"]
        final_rows.append(out)
    return final_rows


def build_final_clean(
    transcriptions: list[dict[str, str]],
    rows_long: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    lookup = {(row["numero"], row["campo_salida"]): row for row in rows_long}

    base_columns = [
        "nombre",
        "edad",
        "genero",
        "numero",
        "departamento",
        "n_educativo",
        "oficio",
        "voto",
        "segmento",
        "voto2",
        "etiqueta",
    ]

    final_rows: list[dict[str, Any]] = []
    for row in transcriptions:
        out = {column: row.get(column) for column in base_columns}
        for question_number in range(1, 13):
            question = f"q{question_number}"
            value = row.get(question)
            if question in OPTION_MAPS:
                value = row.get(f"{question}_texto")
            out[question] = value
            if question in QUESTION_FIELD_ORDER:
                for field in QUESTION_FIELD_ORDER[question]:
                    audit = lookup[(row["numero"], field)]
                    out[field] = audit["codigo_manual_final"]
        final_rows.append(out)
    return final_rows


def build_manual_memo(summary: list[dict[str, Any]], rows_long: list[dict[str, Any]], coherence: list[dict[str, Any]]) -> str:
    manual_states = Counter(row["estado_manual_final"] for row in rows_long)
    sources = Counter(row["fuente_manual_final"] for row in rows_long)
    top_changes = sorted(summary, key=lambda row: row["cambios_manual_vs_revision"], reverse=True)[:10]

    lines = [
        "# Corrección manual focalizada R8 2026",
        "",
        "## Cobertura",
        f"- Decisiones totales: {len(rows_long)}",
        f"- `MANTENER_FINAL`: {manual_states.get('MANTENER_FINAL', 0)}",
        f"- `CORREGIR_FINAL`: {manual_states.get('CORREGIR_FINAL', 0)}",
        f"- `MISSING_VALIDO_FINAL`: {manual_states.get('MISSING_VALIDO_FINAL', 0)}",
        f"- `SIN_ENCAJE_CODEBOOK_FINAL`: {manual_states.get('SIN_ENCAJE_CODEBOOK_FINAL', 0)}",
        "",
        "## Fuente de la decisión final",
        f"- `congelado_mantener`: {sources.get('congelado_mantener', 0)}",
        f"- `conservar_revision_previa`: {sources.get('conservar_revision_previa', 0)}",
        f"- `ausencia_dimension`: {sources.get('ausencia_dimension', 0)}",
        f"- `fallback_ausencia_dimension`: {sources.get('fallback_ausencia_dimension', 0)}",
        f"- `override_explicito`: {sources.get('override_explicito', 0)}",
        "",
        "## Cambios manuales más intensos",
    ]
    for row in top_changes:
        lines.append(
            f"- `{row['pregunta']}/{row['campo_salida']}`: {row['cambios_manual_vs_revision']} cambios manuales vs revisión previa."
        )
    lines.extend(
        [
            "",
            "## Coherencia residual",
            f"- Alertas de coherencia final: {len(coherence)}",
            "- Las alertas restantes reflejan principalmente la decisión de congelar los `MANTENER` previos y no reabrirlos salvo contradicción fuerte.",
        ]
    )
    return "\n".join(lines) + "\n"


def main() -> None:
    audit_rows = load_csv(AUDIT_LONG_INPUT)
    transcriptions = load_csv(TRANSCRIPTION_FILE)

    for row in transcriptions:
        for question in OPTION_MAPS:
            row[f"{question}_texto"] = recode_option_letter(question, row.get(question))

    manual_rows = [manualize_row(row) for row in audit_rows]
    summary_rows = build_summary(manual_rows)
    coherence_rows = build_coherence(manual_rows, transcriptions)
    final_wide_rows = build_final_wide(transcriptions, manual_rows)
    final_clean_rows = build_final_clean(transcriptions, manual_rows)

    write_csv(AUDIT_LONG_OUTPUT, manual_rows, list(manual_rows[0].keys()))

    audit_wb = Workbook()
    audit_wb.remove(audit_wb.active)
    ws_summary = audit_wb.create_sheet("resumen_manual")
    append_sheet(ws_summary, summary_rows, list(summary_rows[0].keys()))
    ws_long = audit_wb.create_sheet("auditoria_manual_long")
    append_sheet(ws_long, manual_rows, list(manual_rows[0].keys()))
    ws_coherence = audit_wb.create_sheet("coherencia_manual")
    coherence_headers = list(coherence_rows[0].keys()) if coherence_rows else ["numero", "par", "nota"]
    append_sheet(ws_coherence, coherence_rows, coherence_headers)
    audit_wb.save(AUDIT_WORKBOOK_OUTPUT)

    final_wb = Workbook()
    ws_final = final_wb.active
    ws_final.title = "R8_codificada_final"
    append_sheet(ws_final, final_wide_rows, list(final_wide_rows[0].keys()))
    final_wb.save(FINAL_WIDE_OUTPUT)

    clean_wb = Workbook()
    ws_clean = clean_wb.active
    ws_clean.title = "R8_codificada"
    append_sheet(ws_clean, final_clean_rows, list(final_clean_rows[0].keys()))
    clean_wb.save(FINAL_CLEAN_OUTPUT)

    write_csv(FINAL_CLEAN_CSV_OUTPUT, final_clean_rows, list(final_clean_rows[0].keys()))

    MANUAL_MEMO_OUTPUT.write_text(
        build_manual_memo(summary_rows, manual_rows, coherence_rows),
        encoding="utf-8",
    )

    print(f"Audit long manual: {AUDIT_LONG_OUTPUT}")
    print(f"Audit workbook manual: {AUDIT_WORKBOOK_OUTPUT}")
    print(f"Final wide workbook: {FINAL_WIDE_OUTPUT}")
    print(f"Final clean workbook: {FINAL_CLEAN_OUTPUT}")
    print(f"Final clean csv: {FINAL_CLEAN_CSV_OUTPUT}")
    print(f"Manual memo: {MANUAL_MEMO_OUTPUT}")


if __name__ == "__main__":
    main()
