from pathlib import Path
from openpyxl import load_workbook

src = Path("/Users/simonherrera/Downloads/clasificacion_R11_2026.xlsx")
repo_root = Path(__file__).resolve().parents[4]
out_dir = repo_root / "data/processed/analysis/2026/R11"
out = out_dir / "R11_codificada.xlsx"

corrections = {
    "codigo_q3": {
        6: "Oferta Técnica y Profesional (UTU)",
        7: None,
        8: None,
        13: None,
        17: None,
        23: None,
        28: None,
        31: None,
        34: None,
        40: "Calidad, Exigencia y Valores",
        41: None,
        43: "Equipos Multidisciplinarios y Apoyo; Actividades Complementarias",
        44: None,
        50: "Calidad, Exigencia y Valores",
        52: None,
        55: None,
        56: None,
        59: None,
        70: None,
        73: "Calidad, Exigencia y Valores",
        80: None,
        81: None,
        83: "Actividades Complementarias",
        84: None,
        85: None,
        92: None,
        93: None,
        97: "Infraestructura y Capacidad; Gestión y Servicios de Apoyo",
        99: "Gestión y Servicios de Apoyo",
        102: "Modalidades Flexibles (Virtualidad)",
        103: "Extensión Horaria; Infraestructura y Capacidad",
        108: "Equipos Multidisciplinarios y Apoyo",
        110: "Infraestructura y Capacidad; Actividades Complementarias",
        117: "Calidad, Exigencia y Valores",
    },
    "codigo_q5": {
        16: "Desconocimiento Total",
        18: "Gobernanza y Participación; Seguridad Social y Jubilaciones; Insumo para Acuerdos Nacionales",
        30: "Seguridad Social y Jubilaciones",
        39: "Gobernanza y Participación",
        40: "Seguridad Social y Jubilaciones; Protección a la Infancia y Transferencias; Burocracia y Escepticismo",
        44: "Protección a la Infancia y Transferencias",
        77: "Desconocimiento Total",
        85: "Desconocimiento Total",
        89: "Desconocimiento Total",
        93: "Seguridad Social y Jubilaciones",
        95: "Seguridad Social y Jubilaciones",
        102: "Insumo para Acuerdos Nacionales",
        107: "Desconocimiento Total",
        115: "Burocracia y Escepticismo",
    },
    "codigo_q6": {
        12: "Garantía y Control Estatal",
        40: "Rechazo por Riesgo de Estatización; Ineficiencia del Estado; Defensa de la Libertad de Elección",
        44: "Defensa de la Libertad de Elección",
        52: "Garantía y Control Estatal",
        54: "Garantía y Control Estatal",
        65: "Desinformación o Indiferencia",
        83: "Propuesta de Equilibrio / Intermedia",
        86: "Garantía y Control Estatal",
        96: "Garantía y Control Estatal",
        99: "Garantía y Control Estatal",
        108: "Garantía y Control Estatal",
    },
    "codigo_q7": {
        15: None,
        16: "Justicia y Foco en Primera Infancia",
        18: "Justicia y Foco en Primera Infancia; Crítica a la Dependencia (Populismo); Necesidad de Control y Monitoreo",
        19: None,
        27: None,
        37: None,
        41: None,
        45: "Justicia y Foco en Primera Infancia",
        46: None,
        58: None,
        61: None,
        71: "Crítica a la Dependencia (Populismo)",
        87: "Justicia y Foco en Primera Infancia",
        89: "Justicia y Foco en Primera Infancia",
        93: "Justicia y Foco en Primera Infancia",
        94: None,
        101: "Justicia y Foco en Primera Infancia",
        102: "Justicia y Foco en Primera Infancia; Percepción de Injusticia / Discriminación",
        103: "Justicia y Foco en Primera Infancia; Percepción de Injusticia / Discriminación",
        105: "Justicia y Foco en Primera Infancia; Necesidad de Control y Monitoreo; Demanda de Soluciones de Fondo (Trabajo)",
        111: "Justicia y Foco en Primera Infancia",
        114: "Justicia y Foco en Primera Infancia",
        119: "Justicia y Foco en Primera Infancia",
    },
    "codigo_q9": {
        4: "Costo de Vida y Salarios; Empleo y Mercado Laboral; Seguridad Ciudadana; Educación y Salud Pública",
        10: None,
        12: None,
        16: "Seguridad Ciudadana",
        37: None,
        53: None,
        65: None,
        76: "Educación y Salud Pública; Seguridad Ciudadana; Costo de Vida y Salarios",
        93: "Cumplimiento de Promesas",
        97: None,
        101: None,
        108: None,
        112: "Revisión de Seguridad Social; Equidad y Género; Costo de Vida y Salarios",
    },
    "codigo_q12": {
        5: None,
        20: "Irregularidades en el Interior; Cuestionamientos a la Ética Pública; Inacción o \"Empate\" Político",
        21: "Caso Cardama y Compra de Buques; Cuestionamientos a la Ética Pública",
        25: None,
        29: "Cuestionamientos a la Ética Pública; Inacción o \"Empate\" Político",
        31: "Cuestionamientos a la Ética Pública",
        36: None,
        39: None,
        47: "Percepción de Persecución Política",
        63: "Percepción de Persecución Política",
        68: None,
        70: "Caso ASSE y Gestión Cipriani; Caso Cardama y Compra de Buques; Irregularidades en el Interior; Cuestionamientos a la Ética Pública",
        76: None,
        80: "Percepción de Persecución Política; Inacción o \"Empate\" Político; Cuestionamientos a la Ética Pública",
        104: None,
        105: "Caso Cardama y Compra de Buques; Caso ASSE y Gestión Cipriani",
        116: "Percepción de Persecución Política; Saturación y Desinterés",
    },
    "codigo_q13": {
        32: None,
        34: "Cuestionamiento por \"Blindaje\" Político",
        35: "Responsabilidad Institucional / Omisión",
        39: "Defensa de la Gestión (Exoneración)",
        47: None,
        51: "Individualización de la Culpa",
        65: "Individualización de la Culpa",
        68: "Cuestionamiento por \"Blindaje\" Político",
        70: "Estilo de Liderazgo Presidencialista; Responsabilidad Institucional / Omisión",
        76: None,
        80: "Defensa de la Gestión (Exoneración); Individualización de la Culpa",
        82: None,
        86: "Cuestionamiento por \"Blindaje\" Político; Responsabilidad Institucional / Omisión",
        96: "Estilo de Liderazgo Presidencialista",
        105: "Responsabilidad Institucional / Omisión",
        116: "Responsabilidad Institucional / Omisión",
    },
}

wb = load_workbook(src)
ws = wb["resultados"]
headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
columns = {h: i + 1 for i, h in enumerate(headers)}

changed = []
for header, rows in corrections.items():
    col = columns[header]
    for row, new_value in rows.items():
        cell = ws.cell(row, col)
        old_value = cell.value
        if old_value != new_value:
            cell.value = new_value
            changed.append((row, header, old_value, new_value))

out_dir.mkdir(parents=True, exist_ok=True)
wb.save(out)

print(f"saved={out}")
print(f"changes={len(changed)}")
for row, header, old, new in changed:
    print(f"{row}\t{header}\t{old!r}\t=>\t{new!r}")
