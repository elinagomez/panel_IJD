import json
from pathlib import Path
from openpyxl import load_workbook

script_dir = Path(__file__).resolve().parent
repo_root = script_dir.parents[3]
definitions = json.load(open(script_dir / "definitions.json", encoding="utf-8"))
path = repo_root / "data/processed/analysis/2026/R11/R11_codificada.xlsx"
wb = load_workbook(path, data_only=True)
ws = wb["resultados"]
headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
columns = {h: i + 1 for i, h in enumerate(headers)}

invalid = []
blank_counts = {}
used_counts = {}
for q, defs in definitions.items():
    allowed = {d["category"] for d in defs}
    used_counts[q] = {}
    blank_counts[q] = 0
    col = columns[f"codigo_{q}"]
    for row in range(2, ws.max_row + 1):
        value = ws.cell(row, col).value
        if value is None or str(value).strip() == "":
            blank_counts[q] += 1
            continue
        for cat in [part.strip() for part in str(value).split(";") if part.strip()]:
            used_counts[q][cat] = used_counts[q].get(cat, 0) + 1
            if cat not in allowed:
                invalid.append((row, q, cat))

print(f"file={path}")
print(f"sheets={wb.sheetnames}")
print(f"dimensions={ws.max_row}x{ws.max_column}")
print(f"invalid_count={len(invalid)}")
print(f"blank_counts={blank_counts}")
print(f"used_counts={json.dumps(used_counts, ensure_ascii=False)}")
if invalid:
    for item in invalid:
        print("INVALID", item)
