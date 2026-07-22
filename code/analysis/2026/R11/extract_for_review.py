import json
from pathlib import Path
from openpyxl import load_workbook

base_path = Path("/Users/simonherrera/Downloads/R11 Base y análisis.xlsx")
class_path = Path("/Users/simonherrera/Downloads/clasificacion_R11_2026.xlsx")
out_dir = Path(__file__).resolve().parent

code_questions = ["q3", "q5", "q6", "q7", "q9", "q12", "q13"]

base_wb = load_workbook(base_path, data_only=True)
definitions = {}
for q in code_questions:
    sheet_name = q if q in base_wb.sheetnames else q.upper()
    ws = base_wb[sheet_name]
    rows = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        if row and row[0] == "Categoría (Etiqueta)":
            start = row
            break
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, 1).value == "Categoría (Etiqueta)":
            for rr in range(r + 1, ws.max_row + 1):
                cat = ws.cell(rr, 1).value
                definition = ws.cell(rr, 2).value
                examples = ws.cell(rr, 3).value
                if cat:
                    rows.append({
                        "category": str(cat),
                        "definition": str(definition or ""),
                        "examples": str(examples or ""),
                    })
            break
    definitions[q] = rows

class_wb = load_workbook(class_path, data_only=True)
ws = class_wb["resultados"]
headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
idx = {h: i + 1 for i, h in enumerate(headers)}

records = []
for r in range(2, ws.max_row + 1):
    rec = {
        "row": r,
        "nombre": ws.cell(r, idx["nombre"]).value,
        "segmento": ws.cell(r, idx["segmento"]).value,
        "voto2": ws.cell(r, idx["voto2"]).value,
    }
    for q in code_questions:
        rec[q] = ws.cell(r, idx[q]).value
        rec[f"codigo_{q}"] = ws.cell(r, idx[f"codigo_{q}"]).value
    records.append(rec)

out_dir.mkdir(parents=True, exist_ok=True)
(out_dir / "definitions.json").write_text(json.dumps(definitions, ensure_ascii=False, indent=2), encoding="utf-8")
(out_dir / "records.json").write_text(json.dumps(records, ensure_ascii=False, indent=2), encoding="utf-8")

print(json.dumps({
    "definition_counts": {k: len(v) for k, v in definitions.items()},
    "records": len(records),
    "headers": headers,
}, ensure_ascii=False, indent=2))
