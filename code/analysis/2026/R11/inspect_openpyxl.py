from pathlib import Path
from openpyxl import load_workbook

files = [
    Path("/Users/simonherrera/Downloads/clasificacion_R11_2026.xlsx"),
    Path("/Users/simonherrera/Downloads/R11 Base y análisis.xlsx"),
]

for path in files:
    print(f"FILE\t{path}")
    wb = load_workbook(path, read_only=False, data_only=False)
    print("SHEETS\t" + "\t".join(wb.sheetnames))
    for ws in wb.worksheets:
        print(f"SHEET\t{ws.title}\t{ws.max_row}\t{ws.max_column}")
        for r in range(1, min(ws.max_row, 8) + 1):
            values = [ws.cell(r, c).value for c in range(1, min(ws.max_column, 20) + 1)]
            print(f"ROW\t{r}\t{values}")
        print()
